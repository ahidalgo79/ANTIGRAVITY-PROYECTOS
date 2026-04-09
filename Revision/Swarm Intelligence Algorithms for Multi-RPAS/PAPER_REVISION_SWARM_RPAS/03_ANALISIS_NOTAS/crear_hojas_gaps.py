import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

ruta = r'C:\Users\HangarUPCH\Documents\Antigravity_Proyectos\Swarm Intelligence Algorithms for Multi-RPAS\PAPER_REVISION_SWARM_RPAS\03_ANALISIS_NOTAS\Fichas_Analisis_NUEVO.xlsx'

print("📊 Creando hojas de GAPS desde Ideas_Gaps.docx...\n")

# === HOJA 1: GAPS_POR_PAPER ===
# Cada fila = un gap específico vinculado a un paper
gaps_por_paper = [
    # PAPER_001 - ACO Coverage Path Planning
    {'ID': 'PAPER_001', 'Dimensión': 'Tecnológica', 'Categoría': 'Seguridad', 'Gap': 'Ausencia de mecanismos de seguridad y evitación de colisiones', 'Prioridad': 'Crítico', 'Naturaleza': 'Omisión explícita'},
    {'ID': 'PAPER_001', 'Dimensión': 'Tecnológica', 'Categoría': 'Hardware', 'Gap': 'Simplificación de la heterogeneidad del hardware (payload, arquitectura)', 'Prioridad': 'Importante', 'Naturaleza': 'Simplificación excesiva'},
    {'ID': 'PAPER_001', 'Dimensión': 'Metodológica', 'Categoría': 'Métricas', 'Gap': 'Omisión del tiempo de retorno a la base', 'Prioridad': 'Importante', 'Naturaleza': 'Omisión reconocida'},
    {'ID': 'PAPER_001', 'Dimensión': 'Metodológica', 'Categoría': 'Validación', 'Gap': 'Validación limitada a entornos de simulación', 'Prioridad': 'Importante', 'Naturaleza': 'Limitación de diseño'},
    {'ID': 'PAPER_001', 'Dimensión': 'Práctica', 'Categoría': 'Operatividad', 'Gap': 'Desatención a restricciones de comunicación y batería', 'Prioridad': 'Crítico', 'Naturaleza': 'Omisión'},
    {'ID': 'PAPER_001', 'Dimensión': 'Práctica', 'Categoría': 'Resiliencia', 'Gap': 'Falta de gestión de fallos de equipo', 'Prioridad': 'Importante', 'Naturaleza': 'Omisión'},
    
    # PAPER_002 - ABC Improved
    {'ID': 'PAPER_002', 'Dimensión': 'Tecnológica', 'Categoría': 'Entornos', 'Gap': 'Limitación en entornos dinámicos (solo estáticos)', 'Prioridad': 'Crítico', 'Naturaleza': 'Omisión reconocida'},
    {'ID': 'PAPER_002', 'Dimensión': 'Metodológica', 'Categoría': 'Validación', 'Gap': 'Ausencia de validación en hardware real', 'Prioridad': 'Importante', 'Naturaleza': 'Simplificación excesiva'},
    {'ID': 'PAPER_002', 'Dimensión': 'Teórica', 'Categoría': 'Escalabilidad', 'Gap': 'Escalabilidad a sistemas multi-UAV (solo un UAV)', 'Prioridad': 'Importante', 'Naturaleza': 'Simplificación'},
    {'ID': 'PAPER_002', 'Dimensión': 'Teórica', 'Categoría': 'Ambiental', 'Gap': 'Omisión de factores ambientales estocásticos (viento, clima)', 'Prioridad': 'Menor', 'Naturaleza': 'Omisión'},
    {'ID': 'PAPER_002', 'Dimensión': 'Metodológica', 'Categoría': 'Parámetros', 'Gap': 'Ajuste empírico de parámetros (depuración manual)', 'Prioridad': 'Menor', 'Naturaleza': 'Limitación no reconocida'},
    
    # PAPER_003 - SSA Modified
    {'ID': 'PAPER_003', 'Dimensión': 'Tecnológica', 'Categoría': 'Escalabilidad', 'Gap': 'Escalabilidad a múltiples agentes (enjambres)', 'Prioridad': 'Importante', 'Naturaleza': 'Omisión'},
    {'ID': 'PAPER_003', 'Dimensión': 'Tecnológica', 'Categoría': 'Hardware', 'Gap': 'Eficiencia computacional frente a hardware embebido', 'Prioridad': 'Importante', 'Naturaleza': 'Limitación no reconocida'},
    {'ID': 'PAPER_003', 'Dimensión': 'Metodológica', 'Categoría': 'Validación', 'Gap': 'Validación limitada a entornos simulados', 'Prioridad': 'Crítico', 'Naturaleza': 'Limitación reconocida'},
    {'ID': 'PAPER_003', 'Dimensión': 'Metodológica', 'Categoría': 'Obstáculos', 'Gap': 'Simplificación del modelado de obstáculos (cilindros)', 'Prioridad': 'Menor', 'Naturaleza': 'Simplificación excesiva'},
    {'ID': 'PAPER_003', 'Dimensión': 'Práctica', 'Categoría': 'Dinámico', 'Gap': 'Ausencia de amenazas y entornos dinámicos', 'Prioridad': 'Crítico', 'Naturaleza': 'Omisión'},
    {'ID': 'PAPER_003', 'Dimensión': 'Práctica', 'Categoría': 'Cinemática', 'Gap': 'Omisión de restricciones cinemáticas avanzadas (velocidad)', 'Prioridad': 'Importante', 'Naturaleza': 'Simplificación excesiva'},
    {'ID': 'PAPER_003', 'Dimensión': 'Teórica', 'Categoría': 'Energía', 'Gap': 'Modelado incompleto de la optimización de energía', 'Prioridad': 'Importante', 'Naturaleza': 'Simplificación excesiva'},
    
    # PAPER_004 - AGA + IABC Multi-UAV
    {'ID': 'PAPER_004', 'Dimensión': 'Tecnológica', 'Categoría': 'Computación', 'Gap': 'Carga computacional y tiempo de procesamiento elevado', 'Prioridad': 'Importante', 'Naturaleza': 'Limitación reconocida'},
    {'ID': 'PAPER_004', 'Dimensión': 'Tecnológica', 'Categoría': 'Tiempo Real', 'Gap': 'Limitaciones de software y tiempo real', 'Prioridad': 'Crítico', 'Naturaleza': 'Limitación no reconocida'},
    {'ID': 'PAPER_004', 'Dimensión': 'Metodológica', 'Categoría': 'Validación', 'Gap': 'Ausencia de pruebas físicas (validación)', 'Prioridad': 'Importante', 'Naturaleza': 'Omisión'},
    {'ID': 'PAPER_004', 'Dimensión': 'Práctica', 'Categoría': 'Ambiental', 'Gap': 'Simplificación de amenazas ambientales (círculos estáticos)', 'Prioridad': 'Crítico', 'Naturaleza': 'Simplificación excesiva'},
    {'ID': 'PAPER_004', 'Dimensión': 'Práctica', 'Categoría': 'Cinemática', 'Gap': 'Omisión de restricciones cinemáticas de los UAVs', 'Prioridad': 'Importante', 'Naturaleza': 'Omisión'},
    {'ID': 'PAPER_004', 'Dimensión': 'Teórica', 'Categoría': 'Dimensionalidad', 'Gap': 'Supuesto de altitud fija (2D vs 3D)', 'Prioridad': 'Importante', 'Naturaleza': 'Simplificación excesiva'},
    
    # PAPER_005 - SSA + BINN
    {'ID': 'PAPER_005', 'Dimensión': 'Tecnológica', 'Categoría': 'Hardware', 'Gap': 'Validación en hardware embebido y limitaciones de cómputo real', 'Prioridad': 'Importante', 'Naturaleza': 'Limitación no reconocida'},
    {'ID': 'PAPER_005', 'Dimensión': 'Tecnológica', 'Categoría': 'Escalabilidad', 'Gap': 'Escalabilidad del enjambre (solo 5 UAVs probados)', 'Prioridad': 'Importante', 'Naturaleza': 'Simplificación excesiva'},
    {'ID': 'PAPER_005', 'Dimensión': 'Metodológica', 'Categoría': 'Validación', 'Gap': 'Ausencia de validación en entornos físicos', 'Prioridad': 'Crítico', 'Naturaleza': 'Omisión reconocida'},
    {'ID': 'PAPER_005', 'Dimensión': 'Metodológica', 'Categoría': 'Métricas', 'Gap': 'Métricas de desempeño incompletas (energía y latencia)', 'Prioridad': 'Menor', 'Naturaleza': 'Omisión'},
    {'ID': 'PAPER_005', 'Dimensión': 'Práctica', 'Categoría': 'Ambiental', 'Gap': 'Omisión de condiciones ambientales adversas', 'Prioridad': 'Importante', 'Naturaleza': 'Omisión'},
    {'ID': 'PAPER_005', 'Dimensión': 'Práctica', 'Categoría': 'Sensores', 'Gap': 'Supuesto de sensores perfectos y comunicación ideal', 'Prioridad': 'Crítico', 'Naturaleza': 'Simplificación excesiva'},
    {'ID': 'PAPER_005', 'Dimensión': 'Teórica', 'Categoría': 'Formación', 'Gap': 'Exclusión del control de formación', 'Prioridad': 'Crítico', 'Naturaleza': 'Simplificación excesiva'},
    
    # PAPER_006 - Swarm Implementation
    {'ID': 'PAPER_006', 'Dimensión': 'Tecnológica', 'Categoría': 'Dinámico', 'Gap': 'Limitación en escalabilidad y manejo de obstáculos dinámicos', 'Prioridad': 'Crítico', 'Naturaleza': 'Simplificación excesiva'},
    {'ID': 'PAPER_006', 'Dimensión': 'Tecnológica', 'Categoría': 'Memoria', 'Gap': 'Dependencia del muestreo para la memoria', 'Prioridad': 'Importante', 'Naturaleza': 'Limitación no reconocida'},
    {'ID': 'PAPER_006', 'Dimensión': 'Metodológica', 'Categoría': 'Métricas', 'Gap': 'Unidimensionalidad de las métricas de optimización (solo brevedad)', 'Prioridad': 'Importante', 'Naturaleza': 'Omisión deliberada'},
    {'ID': 'PAPER_006', 'Dimensión': 'Metodológica', 'Categoría': 'Diseño', 'Gap': 'Ausencia de diversidad en el diseño experimental', 'Prioridad': 'Menor', 'Naturaleza': 'Limitación reconocida'},
    {'ID': 'PAPER_006', 'Dimensión': 'Práctica', 'Categoría': 'Validación', 'Gap': 'Brecha entre simulación y aplicación real', 'Prioridad': 'Crítico', 'Naturaleza': 'Omisión'},
    {'ID': 'PAPER_006', 'Dimensión': 'Práctica', 'Categoría': 'Costo', 'Gap': 'Falta de análisis de costo computacional y energía', 'Prioridad': 'Importante', 'Naturaleza': 'Omisión'},
    
    # PAPER_007 - SIGPAF Fuzzy
    {'ID': 'PAPER_007', 'Dimensión': 'Tecnológica', 'Categoría': 'Computación', 'Gap': 'Gasto computacional de la lógica difusa', 'Prioridad': 'Crítico', 'Naturaleza': 'Limitación reconocida'},
    {'ID': 'PAPER_007', 'Dimensión': 'Tecnológica', 'Categoría': 'Arquitectura', 'Gap': 'Limitación de sistemas MISO en TSK', 'Prioridad': 'Importante', 'Naturaleza': 'Limitación técnica'},
    {'ID': 'PAPER_007', 'Dimensión': 'Metodológica', 'Categoría': 'Aplicaciones', 'Gap': 'Escasez de aplicaciones específicas para USV', 'Prioridad': 'Importante', 'Naturaleza': 'Omisión en literatura'},
    {'ID': 'PAPER_007', 'Dimensión': 'Práctica', 'Categoría': 'Validación', 'Gap': 'Ausencia de pruebas en entornos físicos reales', 'Prioridad': 'Crítico', 'Naturaleza': 'Omisión'},
    {'ID': 'PAPER_007', 'Dimensión': 'Práctica', 'Categoría': 'Hardware', 'Gap': 'Omisión de costos de implementación y hardware', 'Prioridad': 'Importante', 'Naturaleza': 'Omisión total'},
    
    # PAPER_008 - Border Patrol ISFLA
    {'ID': 'PAPER_008', 'Dimensión': 'Tecnológica', 'Categoría': 'Escalabilidad', 'Gap': 'Escalabilidad de la flota y densidad del problema (solo 5 UAVs)', 'Prioridad': 'Importante', 'Naturaleza': 'Simplificación excesiva'},
    {'ID': 'PAPER_008', 'Dimensión': 'Metodológica', 'Categoría': 'Validación', 'Gap': 'Ausencia de validación en entornos físicos reales', 'Prioridad': 'Crítico', 'Naturaleza': 'Omisión'},
    {'ID': 'PAPER_008', 'Dimensión': 'Práctica', 'Categoría': 'Clima', 'Gap': 'Exclusión de condiciones meteorológicas adversas', 'Prioridad': 'Crítico', 'Naturaleza': 'Omisión deliberada'},
    {'ID': 'PAPER_008', 'Dimensión': 'Teórica', 'Categoría': 'Dinámica', 'Gap': 'Simplificación excesiva de la dinámica de vuelo y densidad del aire', 'Prioridad': 'Importante', 'Naturaleza': 'Limitación reconocida'},
    
    # PAPER_009 - SPSO Safety
    {'ID': 'PAPER_009', 'Dimensión': 'Tecnológica', 'Categoría': 'Convergencia', 'Gap': 'Convergencia prematura en óptimos locales', 'Prioridad': 'Importante', 'Naturaleza': 'Limitación reconocida'},
    {'ID': 'PAPER_009', 'Dimensión': 'Tecnológica', 'Categoría': 'Multi-objetivo', 'Gap': 'Escalabilidad ante múltiples objetivos de misión', 'Prioridad': 'Importante', 'Naturaleza': 'Limitación reconocida'},
    {'ID': 'PAPER_009', 'Dimensión': 'Práctica', 'Categoría': 'Dinámico', 'Gap': 'Entornos dinámicos y obstáculos móviles', 'Prioridad': 'Crítico', 'Naturaleza': 'Simplificación excesiva'},
    {'ID': 'PAPER_009', 'Dimensión': 'Práctica', 'Categoría': 'Clima', 'Gap': 'Condiciones ambientales adversas (viento y clima)', 'Prioridad': 'Importante', 'Naturaleza': 'Omisión'},
    {'ID': 'PAPER_009', 'Dimensión': 'Teórica', 'Categoría': 'Restricciones', 'Gap': 'Uso de restricciones duras (hard constraints)', 'Prioridad': 'Crítico', 'Naturaleza': 'Limitación reconocida'},
    {'ID': 'PAPER_009', 'Dimensión': 'Teórica', 'Categoría': 'Multi-UAV', 'Gap': 'Generalización a flotas multi-UAV', 'Prioridad': 'Importante', 'Naturaleza': 'Limitación de arquitectura'},
    
    # PAPER_010 - Review AI Path Planning
    {'ID': 'PAPER_010', 'Dimensión': 'Tecnológica', 'Categoría': 'Escalabilidad', 'Gap': 'Escalabilidad y tasa de fallo de la tarea', 'Prioridad': 'Crítico', 'Naturaleza': 'Limitación reconocida'},
    {'ID': 'PAPER_010', 'Dimensión': 'Tecnológica', 'Categoría': 'Trayectoria', 'Gap': 'Necesidad de suavizado de trayectorias post-algoritmo', 'Prioridad': 'Importante', 'Naturaleza': 'Simplificación excesiva'},
    {'ID': 'PAPER_010', 'Dimensión': 'Tecnológica', 'Categoría': 'Energía', 'Gap': 'Eficiencia energética y dependencia de sensores', 'Prioridad': 'Crítico', 'Naturaleza': 'Limitación tecnológica'},
    {'ID': 'PAPER_010', 'Dimensión': 'Metodológica', 'Categoría': 'Estandarización', 'Gap': 'Falta de estandarización en comunicación de resultados', 'Prioridad': 'Crítico', 'Naturaleza': 'Omisión sistémica'},
    {'ID': 'PAPER_010', 'Dimensión': 'Práctica', 'Categoría': 'Validación', 'Gap': 'Predominio de entornos artificiales sobre mundo real', 'Prioridad': 'Crítico', 'Naturaleza': 'Omisión'},
    {'ID': 'PAPER_010', 'Dimensión': 'Práctica', 'Categoría': 'Comunicación', 'Gap': 'Restricciones en comunicaciones del enjambre', 'Prioridad': 'Importante', 'Naturaleza': 'Simplificación excesiva'},
    
    # PAPER_011 - Review Multi-Target
    {'ID': 'PAPER_011', 'Dimensión': 'Tecnológica', 'Categoría': 'Convergencia', 'Gap': 'Estancamiento en mínimos locales en algoritmos SI', 'Prioridad': 'Crítico', 'Naturaleza': 'Limitación reconocida'},
    {'ID': 'PAPER_011', 'Dimensión': 'Tecnológica', 'Categoría': 'Parámetros', 'Gap': 'Complejidad en ajuste de parámetros de control', 'Prioridad': 'Importante', 'Naturaleza': 'Limitación técnica'},
    {'ID': 'PAPER_011', 'Dimensión': 'Práctica', 'Categoría': 'Tiempo Real', 'Gap': 'Limitación en aplicaciones de tiempo real y entornos dinámicos', 'Prioridad': 'Crítico', 'Naturaleza': 'Limitación de aplicabilidad'},
    {'ID': 'PAPER_011', 'Dimensión': 'Práctica', 'Categoría': 'Energía', 'Gap': 'Omisión del costo energético y de combustible', 'Prioridad': 'Importante', 'Naturaleza': 'Omisión frecuente'},
    {'ID': 'PAPER_011', 'Dimensión': 'Teórica', 'Categoría': 'Objetivos', 'Gap': 'Supuestos de dinámica de objetivos simplificada', 'Prioridad': 'Crítico', 'Naturaleza': 'Simplificación excesiva'},
    
    # PAPER_012 - SPSA Novel
    {'ID': 'PAPER_012', 'Dimensión': 'Tecnológica', 'Categoría': 'Multi-UAV', 'Gap': 'Planificación para múltiples UAVs (swarm planning)', 'Prioridad': 'Importante', 'Naturaleza': 'Omisión'},
    {'ID': 'PAPER_012', 'Dimensión': 'Tecnológica', 'Categoría': 'Dinámico', 'Gap': 'Gestión de obstáculos dinámicos', 'Prioridad': 'Crítico', 'Naturaleza': 'Simplificación excesiva'},
    {'ID': 'PAPER_012', 'Dimensión': 'Metodológica', 'Categoría': 'Validación', 'Gap': 'Validación exclusiva mediante simulación', 'Prioridad': 'Crítico', 'Naturaleza': 'Limitación reconocida'},
    {'ID': 'PAPER_012', 'Dimensión': 'Práctica', 'Categoría': 'Clima', 'Gap': 'Omisión de factores ambientales variables (clima)', 'Prioridad': 'Importante', 'Naturaleza': 'Omisión'},
    {'ID': 'PAPER_012', 'Dimensión': 'Práctica', 'Categoría': 'Dinámica', 'Gap': 'Desatención a la dinámica física del vehículo', 'Prioridad': 'Crítico', 'Naturaleza': 'Simplificación excesiva'},
    {'ID': 'PAPER_012', 'Dimensión': 'Teórica', 'Categoría': 'Dimensionalidad', 'Gap': 'Restricción de movimiento en eje X (reducción dimensional)', 'Prioridad': 'Crítico', 'Naturaleza': 'Simplificación excesiva'},
    
    # PAPER_013 - Chaos PSO
    {'ID': 'PAPER_013', 'Dimensión': 'Tecnológica', 'Categoría': 'Escalabilidad', 'Gap': 'Escalabilidad a sistemas multi-UAV (enjambres)', 'Prioridad': 'Importante', 'Naturaleza': 'Omisión'},
    {'ID': 'PAPER_013', 'Dimensión': 'Metodológica', 'Categoría': 'Validación', 'Gap': 'Falta de validación en entornos físicos', 'Prioridad': 'Crítico', 'Naturaleza': 'Limitación reconocida'},
    {'ID': 'PAPER_013', 'Dimensión': 'Metodológica', 'Categoría': 'Comparación', 'Gap': 'Comparación horizontal con escenarios externos', 'Prioridad': 'Importante', 'Naturaleza': 'Limitación metodológica'},
    {'ID': 'PAPER_013', 'Dimensión': 'Práctica', 'Categoría': 'Dinámico', 'Gap': 'Ausencia de amenazas dinámicas y factores ambientales', 'Prioridad': 'Crítico', 'Naturaleza': 'Simplificación excesiva'},
    {'ID': 'PAPER_013', 'Dimensión': 'Práctica', 'Categoría': 'Restricciones', 'Gap': 'Restricciones físicas limitadas del UAV', 'Prioridad': 'Importante', 'Naturaleza': 'Simplificación excesiva'},
    
    # PAPER_014 - Coverage Path Planning
    {'ID': 'PAPER_014', 'Dimensión': 'Tecnológica', 'Categoría': 'Cinemática', 'Gap': 'Integración de la cinemática del vehículo en planificación', 'Prioridad': 'Crítico', 'Naturaleza': 'Simplificación excesiva'},
    {'ID': 'PAPER_014', 'Dimensión': 'Tecnológica', 'Categoría': 'Viento', 'Gap': 'Estimación y adaptación estocástica del viento en tiempo real', 'Prioridad': 'Importante', 'Naturaleza': 'Omisión y limitación'},
    {'ID': 'PAPER_014', 'Dimensión': 'Metodológica', 'Categoría': 'Framework', 'Gap': 'Carencia de framework multivariable unificado', 'Prioridad': 'Importante', 'Naturaleza': 'Modelo incompleto'},
    {'ID': 'PAPER_014', 'Dimensión': 'Práctica', 'Categoría': 'Comunicación', 'Gap': 'Robustez de las comunicaciones en estrategias multi-UAV', 'Prioridad': 'Crítico', 'Naturaleza': 'Limitación no reconocida'},
    {'ID': 'PAPER_014', 'Dimensión': 'Teórica', 'Categoría': 'UGV vs UAV', 'Gap': 'Brecha conceptual entre métodos terrestres (UGV) y aéreos (UAV)', 'Prioridad': 'Crítico', 'Naturaleza': 'Supuesto no validado'},
    
    # PAPER_015 - Review Optimization Methods
    {'ID': 'PAPER_015', 'Dimensión': 'Tecnológica', 'Categoría': 'Escalabilidad', 'Gap': 'Eficiencia computacional vs escalabilidad en enjambres', 'Prioridad': 'Importante', 'Naturaleza': 'Limitación reconocida'},
    {'ID': 'PAPER_015', 'Dimensión': 'Tecnológica', 'Categoría': 'Comunicación', 'Gap': 'Dependencia de estaciones terrestres (GCS) y comunicación', 'Prioridad': 'Crítico', 'Naturaleza': 'Limitación reconocida'},
    {'ID': 'PAPER_015', 'Dimensión': 'Metodológica', 'Categoría': 'Benchmarks', 'Gap': 'Ausencia de benchmarks y estandarización de datos', 'Prioridad': 'Importante', 'Naturaleza': 'Omisión'},
    {'ID': 'PAPER_015', 'Dimensión': 'Práctica', 'Categoría': 'Validación', 'Gap': 'Brecha entre simulación y realidad (hardware)', 'Prioridad': 'Crítico', 'Naturaleza': 'Omisión sistemática'},
    {'ID': 'PAPER_015', 'Dimensión': 'Práctica', 'Categoría': 'Dinámico', 'Gap': 'Manejo de obstáculos dinámicos y entornos inciertos', 'Prioridad': 'Crítico', 'Naturaleza': 'Simplificación excesiva'},
    {'ID': 'PAPER_015', 'Dimensión': 'Teórica', 'Categoría': 'Ambiental', 'Gap': 'Modelado de factores ambientales externos', 'Prioridad': 'Importante', 'Naturaleza': 'Simplificación excesiva'},
    {'ID': 'PAPER_015', 'Dimensión': 'Teórica', 'Categoría': 'Heterogéneo', 'Gap': 'Generalización en flotas heterogéneas', 'Prioridad': 'Importante', 'Naturaleza': 'Limitación reconocida'},
    
    # PAPER_016 - DDBLPSO
    {'ID': 'PAPER_016', 'Dimensión': 'Tecnológica', 'Categoría': 'Convergencia', 'Gap': 'Inconsistencia en la convergencia (Std elevado)', 'Prioridad': 'Importante', 'Naturaleza': 'Limitación reconocida'},
    {'ID': 'PAPER_016', 'Dimensión': 'Tecnológica', 'Categoría': 'Complejidad', 'Gap': 'Desempeño limitado en funciones de alta complejidad', 'Prioridad': 'Importante', 'Naturaleza': 'Limitación reconocida'},
    {'ID': 'PAPER_016', 'Dimensión': 'Metodológica', 'Categoría': 'Multi-objetivo', 'Gap': 'Simplificación del problema multi-objetivo', 'Prioridad': 'Crítico', 'Naturaleza': 'Simplificación excesiva'},
    {'ID': 'PAPER_016', 'Dimensión': 'Práctica', 'Categoría': 'Hardware', 'Gap': 'Ausencia de validación en hardware real', 'Prioridad': 'Crítico', 'Naturaleza': 'Omisión'},
    {'ID': 'PAPER_016', 'Dimensión': 'Práctica', 'Categoría': 'Estático', 'Gap': 'Entornos estrictamente estáticos', 'Prioridad': 'Importante', 'Naturaleza': 'Simplificación excesiva'},
    
    # PAPER_017 - Survey Optimization
    {'ID': 'PAPER_017', 'Dimensión': 'Tecnológica', 'Categoría': 'Dimensionalidad', 'Gap': 'Escalabilidad en entornos de alta dimensionalidad', 'Prioridad': 'Crítico', 'Naturaleza': 'Limitación reconocida'},
    {'ID': 'PAPER_017', 'Dimensión': 'Tecnológica', 'Categoría': 'ML', 'Gap': 'Costo computacional del aprendizaje automático para tiempo real', 'Prioridad': 'Importante', 'Naturaleza': 'Limitación no reconocida'},
    {'ID': 'PAPER_017', 'Dimensión': 'Metodológica', 'Categoría': 'Métricas', 'Gap': 'Desequilibrio en evaluación de métricas (costo vs calidad)', 'Prioridad': 'Menor', 'Naturaleza': 'Simplificación excesiva'},
    {'ID': 'PAPER_017', 'Dimensión': 'Metodológica', 'Categoría': 'Estático', 'Gap': 'Predominio de validación en entornos estáticos', 'Prioridad': 'Crítico', 'Naturaleza': 'Simplificación excesiva'},
    {'ID': 'PAPER_017', 'Dimensión': 'Práctica', 'Categoría': 'Incertidumbre', 'Gap': 'Omisión de la incertidumbre ambiental', 'Prioridad': 'Crítico', 'Naturaleza': 'Omisión'},
    {'ID': 'PAPER_017', 'Dimensión': 'Práctica', 'Categoría': 'QoS', 'Gap': 'Falta de integración de calidad de servicio (QoS)', 'Prioridad': 'Importante', 'Naturaleza': 'Omisión'},
    {'ID': 'PAPER_017', 'Dimensión': 'Teórica', 'Categoría': 'Información', 'Gap': 'Dependencia de información ambiental completa', 'Prioridad': 'Importante', 'Naturaleza': 'Limitación reconocida'},
    {'ID': 'PAPER_017', 'Dimensión': 'Teórica', 'Categoría': 'Convergencia', 'Gap': 'Ausencia de garantías teóricas de convergencia en meta-heurísticas', 'Prioridad': 'Importante', 'Naturaleza': 'Limitación intrínseca'},
    
    # PAPER_018 - ANFIS-BCO
    {'ID': 'PAPER_018', 'Dimensión': 'Tecnológica', 'Categoría': 'Dimensionalidad', 'Gap': 'Escalabilidad y dimensionalidad del control (2D vs 3D)', 'Prioridad': 'Importante', 'Naturaleza': 'Simplificación excesiva'},
    {'ID': 'PAPER_018', 'Dimensión': 'Metodológica', 'Categoría': 'Robustez', 'Gap': 'Ausencia de pruebas de robustez ante perturbaciones y ruido', 'Prioridad': 'Crítico', 'Naturaleza': 'Omisión'},
    {'ID': 'PAPER_018', 'Dimensión': 'Práctica', 'Categoría': 'Hardware', 'Gap': 'Falta de validación en hardware real', 'Prioridad': 'Crítico', 'Naturaleza': 'Limitación reconocida'},
    {'ID': 'PAPER_018', 'Dimensión': 'Teórica', 'Categoría': 'Linealización', 'Gap': 'Dependencia de modelos linealizados para sistemas no lineales', 'Prioridad': 'Importante', 'Naturaleza': 'Simplificación excesiva'},
    
    # PAPER_019 - MMACO Convergence
    {'ID': 'PAPER_019', 'Dimensión': 'Tecnológica', 'Categoría': 'Hardware', 'Gap': 'Ausencia de despliegue en hardware físico', 'Prioridad': 'Crítico', 'Naturaleza': 'Limitación reconocida'},
    {'ID': 'PAPER_019', 'Dimensión': 'Tecnológica', 'Categoría': 'Tiempo Real', 'Gap': 'Dependencia de cálculos previos (falta de respuesta en tiempo real)', 'Prioridad': 'Crítico', 'Naturaleza': 'Limitación no resuelta'},
    {'ID': 'PAPER_019', 'Dimensión': 'Metodológica', 'Categoría': 'Escalabilidad', 'Gap': 'Limitación en escalabilidad de la flota (solo 2 UAVs)', 'Prioridad': 'Importante', 'Naturaleza': 'Simplificación excesiva'},
    {'ID': 'PAPER_019', 'Dimensión': 'Metodológica', 'Categoría': 'Métricas', 'Gap': 'Falta de métricas de consumo energético y latencia', 'Prioridad': 'Importante', 'Naturaleza': 'Omisión'},
    {'ID': 'PAPER_019', 'Dimensión': 'Práctica', 'Categoría': 'Amenazas', 'Gap': 'Modelado estático de amenazas dinámicas', 'Prioridad': 'Importante', 'Naturaleza': 'Simplificación'},
    {'ID': 'PAPER_019', 'Dimensión': 'Teórica', 'Categoría': 'Dimensionalidad', 'Gap': 'Discretización del espacio 3D en planos 2D', 'Prioridad': 'Importante', 'Naturaleza': 'Simplificación excesiva'},
    {'ID': 'PAPER_019', 'Dimensión': 'Teórica', 'Categoría': 'Información', 'Gap': 'Supuesto de información completa del entorno', 'Prioridad': 'Crítico', 'Naturaleza': 'Simplificación'},
]

# Crear DataFrame
df_gaps_paper = pd.DataFrame(gaps_por_paper)

# === HOJA 2: GAPS_AGRUPADOS ===
# Conteo por dimensión, categoría y prioridad
gaps_agrupados_data = []

# Por Dimensión
for dim in df_gaps_paper['Dimensión'].unique():
    count = len(df_gaps_paper[df_gaps_paper['Dimensión'] == dim])
    gaps_agrupados_data.append(['Dimensión', dim, count, f'{count/len(df_gaps_paper)*100:.1f}%'])

# Por Categoría
for cat in df_gaps_paper['Categoría'].unique():
    count = len(df_gaps_paper[df_gaps_paper['Categoría'] == cat])
    gaps_agrupados_data.append(['Categoría', cat, count, f'{count/len(df_gaps_paper)*100:.1f}%'])

# Por Prioridad
for pri in df_gaps_paper['Prioridad'].unique():
    count = len(df_gaps_paper[df_gaps_paper['Prioridad'] == pri])
    gaps_agrupados_data.append(['Prioridad', pri, count, f'{count/len(df_gaps_paper)*100:.1f}%'])

# Por Naturaleza
for nat in df_gaps_paper['Naturaleza'].unique():
    count = len(df_gaps_paper[df_gaps_paper['Naturaleza'] == nat])
    gaps_agrupados_data.append(['Naturaleza', nat, count, f'{count/len(df_gaps_paper)*100:.1f}%'])

df_gaps_agrupados = pd.DataFrame(gaps_agrupados_data, columns=['Tipo', 'Nombre', 'Cantidad', 'Porcentaje'])

# === HOJA 3: OPORTUNIDADES_TESIS ===
# Priorización de gaps para investigación doctoral
oportunidades_tesis = [
    ['OPORTUNIDADES PARA TESIS DOCTORAL', '', '', ''],
    ['', '', '', ''],
    ['PRIORIDAD', 'OPORTUNIDAD', 'GAPS RELACIONADOS', 'PAPERS AFECTADOS'],
    ['1', 'Estandarización de métricas', 'Falta de benchmarks, métricas incompletas, sin estandarización', '25+ papers'],
    ['2', 'Validación experimental con UAVs reales', 'Ausencia de validación en hardware, solo simulación', '28 papers'],
    ['3', 'Entornos dinámicos y obstáculos móviles', 'Solo entornos estáticos, sin replanificación en tiempo real', '22 papers'],
    ['4', 'Métricas avanzadas (robustez, escalabilidad, tolerancia a fallos)', 'No evaluadas en literatura', '25+ papers'],
    ['5', 'Comparación sistemática (mismos parámetros, mismos escenarios)', 'Cada paper usa diferentes benchmarks', '20+ papers'],
    ['6', 'Factores ambientales (viento, clima, densidad del aire)', 'Omitidos o simplificados', '18 papers'],
    ['7', 'Multi-UAV y coordinación de enjambres', 'Mayoría son single-UAV o flotas pequeñas', '20 papers'],
    ['8', 'Restricciones cinemáticas realistas', 'Simplificación de dinámica de vuelo', '15 papers'],
    ['', '', '', ''],
    ['RECOMENDACIÓN PRINCIPAL', '', '', ''],
    ['Tu tesis puede abordar los gaps 1, 2 y 3 como contribución principal:', '', '', ''],
    ['- Crear un benchmark estandarizado para comparación de algoritmos', '', '', ''],
    ['- Implementar validación experimental con UAVs reales en entorno controlado', '', '', ''],
    ['- Integrar entornos dinámicos con obstáculos móviles y replanificación en tiempo real', '', '', ''],
    ['', '', '', ''],
    ['Fecha de análisis:', datetime.now().strftime('%Y-%m-%d %H:%M'), '', ''],
]

df_oportunidades = pd.DataFrame(oportunidades_tesis, columns=['PRIORIDAD', 'OPORTUNIDAD', 'GAPS RELACIONADOS', 'PAPERS AFECTADOS'])

# === GUARDAR EN EXCEL ===
wb = load_workbook(ruta)

# Eliminar hojas si existen
for sheet in ['GAPS_POR_PAPER', 'GAPS_AGRUPADOS', 'OPORTUNIDADES_TESIS']:
    if sheet in wb.sheetnames:
        del wb[sheet]
wb.save(ruta)

with pd.ExcelWriter(ruta, engine='openpyxl', mode='a') as writer:
    df_gaps_paper.to_excel(writer, sheet_name='GAPS_POR_PAPER', index=False)
    df_gaps_agrupados.to_excel(writer, sheet_name='GAPS_AGRUPADOS', index=False)
    df_oportunidades.to_excel(writer, sheet_name='OPORTUNIDADES_TESIS', index=False)

# === MOSTRAR RESULTADOS ===
print(f"{'='*60}")
print(f"✅ HOJAS DE GAPS CREADAS EXITOSAMENTE")
print(f"{'='*60}")
print(f"\nGAPS_POR_PAPER: {len(df_gaps_paper)} gaps identificados")
print(f"\nDistribución por Dimensión:")
print(df_gaps_paper['Dimensión'].value_counts())
print(f"\nDistribución por Prioridad:")
print(df_gaps_paper['Prioridad'].value_counts())
print(f"\nHojas creadas:")
print(f"  - GAPS_POR_PAPER ({len(df_gaps_paper)} filas)")
print(f"  - GAPS_AGRUPADOS ({len(df_gaps_agrupados)} filas)")
print(f"  - OPORTUNIDADES_TESIS (16 filas)")