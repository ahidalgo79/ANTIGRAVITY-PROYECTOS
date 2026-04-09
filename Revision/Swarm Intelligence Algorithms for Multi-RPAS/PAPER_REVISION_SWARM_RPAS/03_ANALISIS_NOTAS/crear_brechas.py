import pandas as pd
from openpyxl import load_workbook

ruta = r'C:\Users\HangarUPCH\Documents\Antigravity_Proyectos\Swarm Intelligence Algorithms for Multi-RPAS\PAPER_REVISION_SWARM_RPAS\03_ANALISIS_NOTAS\Fichas_Analisis_NUEVO.xlsx'

# Cargar datos
df_metricas = pd.read_excel(ruta, sheet_name='METRICAS_COMPARATIVAS')
df_todos = pd.read_excel(ruta, sheet_name='TODOS_PAPERS')

print("🔍 Analizando brechas de investigación...\n")

# === BRECHA 1: MÉTRICAS NO REPORTADAS ===
brechas_metricas = []

for idx, row in df_metricas.iterrows():
    paper_id = row['ID']
    algoritmo = row['Algoritmo']
    
    # Verificar cada métrica
    if pd.isna(row['Tiempo (num)']) and ('No especificada' in str(row['Tiempo (texto)']) or row['Tiempo (texto)'] == ''):
        brechas_metricas.append({
            'ID': paper_id,
            'Algoritmo': algoritmo,
            'Tipo Brecha': 'Métrica No Reportada',
            'Categoría': 'Tiempo',
            'Descripción': 'No se reportan valores de tiempo de ejecución',
            'Recomendación': 'Incluir tiempo de cómputo en segundos/milisegundos'
        })
    
    if pd.isna(row['Energía (num)']) and ('No especificada' in str(row['Energía (texto)']) or row['Energía (texto)'] == ''):
        brechas_metricas.append({
            'ID': paper_id,
            'Algoritmo': algoritmo,
            'Tipo Brecha': 'Métrica No Reportada',
            'Categoría': 'Energía',
            'Descripción': 'No se reportan valores de consumo energético',
            'Recomendación': 'Medir consumo en Joules, % de batería, o distancia equivalente'
        })
    
    if pd.isna(row['Convergencia (num)']) and ('No especificada' in str(row['Convergencia (texto)']) or row['Convergencia (texto)'] == ''):
        brechas_metricas.append({
            'ID': paper_id,
            'Algoritmo': algoritmo,
            'Tipo Brecha': 'Métrica No Reportada',
            'Categoría': 'Convergencia',
            'Descripción': 'No se reportan criterios de convergencia',
            'Recomendación': 'Especificar iteraciones máximas, tolerancia, o criterio de parada'
        })

# === BRECHA 2: ALGORITMOS SIN BENCHMARK ===
# Algoritmos que no comparan con otros métodos
algorithms_no_benchmark = []
for idx, row in df_todos.iterrows():
    if row['Algoritmo Principal'] != 'Review':
        abstract = str(row.get('Abstract Resumido', ''))
        comparacion = str(row.get('Comparación Con', ''))
        
        if 'comparación' not in abstract.lower() and 'vs' not in abstract.lower() and 'benchmark' not in abstract.lower():
            algorithms_no_benchmark.append({
                'ID': row['ID'],
                'Algoritmo': row['Algoritmo Principal'],
                'Tipo Brecha': 'Sin Benchmark',
                'Categoría': 'Validación',
                'Descripción': 'No compara con algoritmos de referencia',
                'Recomendación': 'Comparar con PSO, ACO, GA u otros algoritmos estándar'
            })

# === BRECHA 3: FALTA DE VALIDACIÓN EXPERIMENTAL ===
brechas_validacion = []
for idx, row in df_todos.iterrows():
    validacion = str(row.get('Validación', ''))
    pruebas_reales = str(row.get('Pruebas Reales', ''))
    
    if 'simulación' in validacion.lower() and 'real' not in pruebas_reales.lower():
        brechas_validacion.append({
            'ID': row['ID'],
            'Algoritmo': row['Algoritmo Principal'],
            'Tipo Brecha': 'Sin Validación Real',
            'Categoría': 'Validación',
            'Descripción': 'Solo validación en simulación, sin pruebas reales',
            'Recomendación': 'Incluir experimentos con UAVs reales en entorno controlado'
        })

# === BRECHA 4: MÉTRICAS AVANZADAS AUSENTES ===
metricas_avanzadas = ['Robustez', 'Escalabilidad', 'Tiempo Real', 'Tolerancia a Fallos']
brechas_avanzadas = []

for idx, row in df_todos.iterrows():
    if row['Algoritmo Principal'] != 'Review':
        texto_completo = str(row.get('Abstract Resumido', '')) + str(row.get('Notas Personales', ''))
        
        for metrica in metricas_avanzadas:
            if metrica.lower() not in texto_completo.lower():
                brechas_avanzadas.append({
                    'ID': row['ID'],
                    'Algoritmo': row['Algoritmo Principal'],
                    'Tipo Brecha': 'Métrica Avanzada Ausente',
                    'Categoría': metrica,
                    'Descripción': f'No evalúa {metrica.lower()}',
                    'Recomendación': f'Incluir evaluación de {metrica.lower()} en experimentos futuros'
                })

# === COMBINAR TODAS LAS BRECHAS ===
todas_las_brechas = brechas_metricas + algorithms_no_benchmark + brechas_validacion + brechas_avanzadas
df_brechas = pd.DataFrame(todas_las_brechas)

# === RESUMEN ESTADÍSTICO ===
resumen_data = [
    ['RESUMEN DE BRECHAS IDENTIFICADAS', ''],
    ['Total de brechas detectadas:', len(df_brechas)],
    ['', ''],
    ['Por categoría:', ''],
]

for categoria, count in df_brechas['Categoría'].value_counts().items():
    resumen_data.append([f'  {categoria}:', count])

resumen_data.extend([
    ['', ''],
    ['Por tipo de brecha:', ''],
])

for tipo, count in df_brechas['Tipo Brecha'].value_counts().items():
    resumen_data.append([f'  {tipo}:', count])

resumen_data.extend([
    ['', ''],
    ['OPORTUNIDADES PARA TESIS DOCTORAL:', ''],
    ['1. Estandarización de métricas', 'Crear benchmark común para comparación de algoritmos'],
    ['2. Validación experimental', 'Implementar pruebas con UAVs reales'],
    ['3. Métricas avanzadas', 'Evaluar robustez, escalabilidad y tolerancia a fallos'],
    ['4. Comparación sistemática', 'Evaluar todos los algoritmos con mismos parámetros'],
    ['', ''],
    ['Fecha de análisis:', pd.Timestamp.now().strftime('%Y-%m-%d %H:%M')],
])

df_resumen = pd.DataFrame(resumen_data, columns=['BRECHAS DE INVESTIGACIÓN', 'Valor'])

# === GUARDAR EN EXCEL ===
wb = load_workbook(ruta)

# Eliminar hoja si existe
if 'BRECHAS_INVESTIGACION' in wb.sheetnames:
    del wb['BRECHAS_INVESTIGACION']
wb.save(ruta)

with pd.ExcelWriter(ruta, engine='openpyxl', mode='a') as writer:
    df_resumen.to_excel(writer, sheet_name='BRECHAS_INVESTIGACION', index=False)
    df_brechas.to_excel(writer, sheet_name='BRECHAS_DETALLADAS', index=False)

# === MOSTRAR RESULTADOS ===
print(f"{'='*60}")
print(f"✅ BRECHAS DE INVESTIGACIÓN IDENTIFICADAS")
print(f"{'='*60}")
print(f"\nTotal de brechas: {len(df_brechas)}")
print(f"\nPor categoría:")
print(df_brechas['Categoría'].value_counts())
print(f"\nPor tipo:")
print(df_brechas['Tipo Brecha'].value_counts())
print(f"\nHojas creadas:")
print(f"  - BRECHAS_INVESTIGACION (resumen)")
print(f"  - BRECHAS_DETALLADAS (lista completa)")