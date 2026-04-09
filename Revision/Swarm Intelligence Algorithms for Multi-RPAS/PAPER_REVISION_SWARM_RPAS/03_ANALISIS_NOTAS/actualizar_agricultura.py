import pandas as pd
from openpyxl import load_workbook

ruta = r'C:\Users\HangarUPCH\Documents\Antigravity_Proyectos\Swarm Intelligence Algorithms for Multi-RPAS\PAPER_REVISION_SWARM_RPAS\03_ANALISIS_NOTAS\Fichas_Analisis_NUEVO.xlsx'

# Cargar TODOS_PAPERS
df = pd.read_excel(ruta, sheet_name='TODOS_PAPERS')

# Filtrar por Agricultura (varias formas de escribirlo)
agricultura = df[df['Aplicación Principal'].str.contains('Agricultura|Agri|Agrícola', case=False, na=False)].copy()

# Eliminar hoja AGRICULTURA existente
wb = load_workbook(ruta)
if 'AGRICULTURA' in wb.sheetnames:
    del wb['AGRICULTURA']
wb.save(ruta)

# Guardar nueva hoja AGRICULTURA
with pd.ExcelWriter(ruta, engine='openpyxl', mode='a') as writer:
    agricultura.to_excel(writer, sheet_name='AGRICULTURA', index=False)

print(f"✅ AGRICULTURA actualizado: {len(agricultura)} papers")
print(f"IDs: {agricultura['ID'].tolist()}")