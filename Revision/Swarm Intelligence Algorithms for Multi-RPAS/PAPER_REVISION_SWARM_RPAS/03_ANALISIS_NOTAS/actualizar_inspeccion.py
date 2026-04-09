import pandas as pd
from openpyxl import load_workbook

ruta = r'C:\Users\HangarUPCH\Documents\Antigravity_Proyectos\Swarm Intelligence Algorithms for Multi-RPAS\PAPER_REVISION_SWARM_RPAS\03_ANALISIS_NOTAS\Fichas_Analisis_NUEVO.xlsx'

# Cargar TODOS_PAPERS
df = pd.read_excel(ruta, sheet_name='TODOS_PAPERS')

# Filtrar por Inspección/Vigilancia/Monitoreo
inspeccion = df[df['Aplicación Principal'].str.contains('Inspeccion|Vigilancia|Monitor', case=False, na=False)].copy()

# También buscar en el título por si acaso
if len(inspeccion) == 0:
    inspeccion = df[df['Título Completo'].str.contains('Inspect|Monitor|Vigil', case=False, na=False)].copy()

# Eliminar hoja INSPECCION existente
wb = load_workbook(ruta)
if 'INSPECCION' in wb.sheetnames:
    del wb['INSPECCION']
wb.save(ruta)

# Guardar nueva hoja INSPECCION
with pd.ExcelWriter(ruta, engine='openpyxl', mode='a') as writer:
    inspeccion.to_excel(writer, sheet_name='INSPECCION', index=False)

print(f"✅ INSPECCION actualizado: {len(inspeccion)} papers")
print(f"IDs: {inspeccion['ID'].tolist() if len(inspeccion) > 0 else 'Sin registros'}")