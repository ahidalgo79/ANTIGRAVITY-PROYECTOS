import openpyxl

file_path = r'C:\Users\HangarUPCH\Documents\Antigravity_Proyectos\Swarm Intelligence Algorithms for Multi-RPAS\PAPER_REVISION_SWARM_RPAS\03_ANALISIS_NOTAS\extraccion_S34_S35_S36.xlsx'
wb = openpyxl.load_workbook(file_path)

missing_data_common = {
    'I1: Publicado': 'Sí (Confirmado)',
    'I2: Peer-reviewed': 'Sí',
    'I3: Texto en inglés': 'Sí',
    'E4: ¿Excluir por método no-SI?': 'No (Aplica SI)',
    'E5: ¿Excluir por dominio no-agrícola?': 'No (Es agrícola)',
    'Notas MMAT': 'Evaluación basada en resumen del texto completo.',
    'Acción requerida en manuscrito': 'Mover a la lista de estudios EXCLUIDOS a texto completo por criterio E3.'
}

specific_data = {
    'S34': {'Número total de gaps identificados': '4'},
    'S35': {'Número total de gaps identificados': '3'},
    'S36': {'Número total de gaps identificados': '4'}
}

for sheet_name in ['S34', 'S35', 'S36']:
    ws = wb[sheet_name]
    sheet_data = {**missing_data_common, **specific_data[sheet_name]}
    
    for row in range(1, ws.max_row + 1):
        cell_label = ws.cell(row=row, column=2).value
        if cell_label:
            cell_str = str(cell_label).strip()
            for key, val in sheet_data.items():
                if cell_str.startswith(key):
                    ws.cell(row=row, column=3).value = val
                    break

wb.save(file_path)
print("Información faltante completada.")
