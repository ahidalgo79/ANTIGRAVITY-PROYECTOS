import openpyxl

file_path = r'C:\Users\HangarUPCH\Documents\Antigravity_Proyectos\Swarm Intelligence Algorithms for Multi-RPAS\PAPER_REVISION_SWARM_RPAS\03_ANALISIS_NOTAS\extraccion_S34_S35_S36.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['S36']

data = {
    'I4:': 'Sí; emplea versión mejorada de ACO (Inteligencia de Enjambres).',
    'I4b:': '1 (single-agent).',
    'I5:': 'Sí; navegación autónoma en huertos para monitoreo, recolección y fumigación.',
    'E3:': 'Sí (Dado que es single-UAV en simulación, se activa la exclusión E3).',
    'DECISIÓN FINAL': 'EXCLUIR [E3] (Nota: modelo Single-UAV, incumple el requisito de enjambre multi-UAV)',
    
    'Tipo de algoritmo SI': 'ACO (Híbrido con teoría de liderazgo)',
    'Nombre exacto del algoritmo': 'Leader Ant Optimization (LAO) algorithm',
    'Modificaciones al algoritmo base': 'Teoría del líder (bellwether) para selección de rutas; ajusta feromonas y heurística con parámetros 3D.',
    'Tamaño del enjambre': '50',
    'Número de UAVs en el framework': '1',
    'Tipo de validación': 'Solo simulación',
    'Entorno de simulación utilizado': 'MATLAB R2020a',
    '¿Incorpora factores ambientales dinámicos?': 'No (simulación estática de huerto, no modelan viento/lluvia)',
    'Tipo de aplicación agrícola': 'Monitoreo, fumigación y asistencia en cosecha',
    'Parámetros agrícolas específicos': 'Dimensiones de árboles frutales (5-10m), cercas (1.2-5m).',
    
    'Tiempo de ejecución / convergencia': 'Sí (0.1021 s promedio)',
    'Consumo de energía / batería': 'No reporta (solo distancia total: 53.74m)',
    'Criterio de convergencia': 'Máximo 100 iteraciones',
    'Comparación con algoritmo baseline': 'Sí (GA, PSO, ACO clásico)',
    'Otras métricas reportadas': 'Fitness medio (44.84), medias de convergencia y p-values.',
    
    'C1 —': '0.0',
    'C2 —': '0.67',
    'C3 —': '1.0',
    'C4 —': '1.0',
    'Q =': '0.5325',
    'Clasificación DRS': 'Moderate',
    
    'Q1:': 'Y',
    'Q2:': 'Y',
    'Q3:': 'Y',
    'Q4:': 'Y',
    'Q5:': 'Y',
    'Número de Y': '5',
    'Clasificación MMAT': 'High',
    
    'Gaps Tecnológicos identificados': 'Necesidad de integración con percepción visual 3D para tiempo real.',
    'Gaps Prácticos identificados': 'Dificultad de mantener ruta óptima ante cambios dinámicos rápidos.',
    'Gaps Metodológicos identificados': 'Falta de comparación con métodos estado-del-arte (SOTA) en 3D.',
    'Gaps Teóricos identificados': 'Selecciones pueden ser subóptimas ante cambios no entrenados.',
    
    'Estado de extracción': 'COMPLETO',
    'Fecha de extracción': '29/03/2026',
    'Revisor': 'Antigravity / NotebookLM',
    'Notas adicionales': 'Igual que S34 y S35, el documento se descarta porque el framework simulado es para un solo dron (1 UAV), a pesar de que NotebookLM pusiera "INCLUIR".'
}

for row in range(1, ws.max_row + 1):
    cell_a = ws.cell(row=row, column=1).value
    if cell_a:
        cell_str = str(cell_a).strip()
        for key, val in data.items():
            if cell_str.startswith(key):
                ws.cell(row=row, column=2).value = val
                break

wb.save(file_path)
print("S36 guardado con exito.")
