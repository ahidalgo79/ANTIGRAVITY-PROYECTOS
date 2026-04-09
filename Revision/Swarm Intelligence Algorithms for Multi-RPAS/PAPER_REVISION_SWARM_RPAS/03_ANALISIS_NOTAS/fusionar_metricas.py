import pandas as pd
from openpyxl import load_workbook

ruta = r'C:\Users\HangarUPCH\Documents\Antigravity_Proyectos\Swarm Intelligence Algorithms for Multi-RPAS\PAPER_REVISION_SWARM_RPAS\03_ANALISIS_NOTAS\Fichas_Analisis_NUEVO.xlsx'

# Cargar ambas hojas
df_metricas = pd.read_excel(ruta, sheet_name='METRICAS_COMPARATIVAS')
df_nuevas = pd.read_excel(ruta, sheet_name='METRICAS_021_034')

print("🔀 Fusionando métricas de papers 021-034...\n")

# Contar actualizaciones
actualizaciones = 0

for idx, row_nueva in df_nuevas.iterrows():
    paper_id = row_nueva['ID']
    
    # Buscar índice en df_metricas
    idx_match = df_metricas[df_metricas['ID'] == paper_id].index
    if len(idx_match) == 0:
        print(f"⚠️ {paper_id}: No encontrado en METRICAS_COMPARATIVAS")
        continue
    idx_match = idx_match[0]
    
    # Actualizar solo si hay valores nuevos (no NaN)
    if pd.notna(row_nueva.get('Tiempo (num)')):
        df_metricas.at[idx_match, 'Tiempo (num)'] = row_nueva['Tiempo (num)']
    
    if pd.notna(row_nueva.get('Energía (num)')):
        df_metricas.at[idx_match, 'Energía (num)'] = row_nueva['Energía (num)']
    
    if pd.notna(row_nueva.get('Convergencia (num)')):
        df_metricas.at[idx_match, 'Convergencia (num)'] = row_nueva['Convergencia (num)']
    
    # Actualizar columnas de texto con lo encontrado
    if row_nueva.get('Tiempo (encontrado)', '') and row_nueva['Tiempo (encontrado)'] != 'No encontrado':
        df_metricas.at[idx_match, 'Tiempo (texto)'] = row_nueva['Tiempo (encontrado)'][:100]
    
    if row_nueva.get('Energía (encontrado)', '') and row_nueva['Energía (encontrado)'] != 'No encontrado':
        df_metricas.at[idx_match, 'Energía (texto)'] = row_nueva['Energía (encontrado)'][:100]
    
    if row_nueva.get('Convergencia (encontrado)', '') and row_nueva['Convergencia (encontrado)'] != 'No encontrado':
        df_metricas.at[idx_match, 'Convergencia (texto)'] = row_nueva['Convergencia (encontrado)'][:100]
    
    actualizaciones += 1
    print(f"✅ {paper_id}: Actualizado")

# Guardar hoja actualizada
wb = load_workbook(ruta)
if 'METRICAS_COMPARATIVAS' in wb.sheetnames:
    del wb['METRICAS_COMPARATIVAS']
wb.save(ruta)

with pd.ExcelWriter(ruta, engine='openpyxl', mode='a') as writer:
    df_metricas.to_excel(writer, sheet_name='METRICAS_COMPARATIVAS', index=False)

# Mostrar resumen final
print(f"\n{'='*60}")
print(f"✅ FUSIÓN COMPLETADA")
print(f"{'='*60}")
print(f"Papers actualizados: {actualizaciones}")

# Contar valores numéricos totales
tiempo_count = df_metricas['Tiempo (num)'].notna().sum()
energia_count = df_metricas['Energía (num)'].notna().sum()
convergencia_count = df_metricas['Convergencia (num)'].notna().sum()

print(f"\nValores numéricos en METRICAS_COMPARATIVAS:")
print(f"  - Tiempo: {tiempo_count}/33 ({tiempo_count/33*100:.1f}%)")
print(f"  - Energía: {energia_count}/33 ({energia_count/33*100:.1f}%)")
print(f"  - Convergencia: {convergencia_count}/33 ({convergencia_count/33*100:.1f}%)")