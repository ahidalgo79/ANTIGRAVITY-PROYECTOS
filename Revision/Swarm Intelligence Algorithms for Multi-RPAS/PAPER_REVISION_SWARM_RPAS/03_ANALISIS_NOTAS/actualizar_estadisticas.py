import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

ruta = r'C:\Users\HangarUPCH\Documents\Antigravity_Proyectos\Swarm Intelligence Algorithms for Multi-RPAS\PAPER_REVISION_SWARM_RPAS\03_ANALISIS_NOTAS\Fichas_Analisis_NUEVO.xlsx'

# Cargar TODOS_PAPERS
df = pd.read_excel(ruta, sheet_name='TODOS_PAPERS')

# === CALCULAR ESTADÍSTICAS ===
total_papers = len(df)
analizados = len(df[df['Estado Lectura'] != 'Pendiente'])
pendientes = len(df[df['Estado Lectura'] == 'Pendiente'])
progreso = (analizados / total_papers * 100) if total_papers > 0 else 0

# Por algoritmo
algo_counts = df['Algoritmo Principal'].value_counts()

# Por año
year_counts = df['Año'].value_counts().sort_index()

# Por tipo de publicación
tipo_counts = df['Tipo Publicación'].value_counts()

# Por validación
val_counts = df['Validación'].value_counts()

# Por relevancia
rel_counts = df['Relevancia'].value_counts()

# Por aplicación principal
app_counts = df['Aplicación Principal'].value_counts()

# === CREAR DATAFRAME DE ESTADÍSTICAS ===
stats_data = [
    ['ESTADÍSTICAS GENERALES DEL ANÁLISIS', ''],
    ['', ''],
    ['Total de Papers:', total_papers],
    ['Papers Analizados:', analizados],
    ['Papers Pendientes:', pendientes],
    ['% Progreso:', f'{progreso:.1f}%'],
    ['', ''],
    ['DISTRIBUCIÓN POR ALGORITMO', ''],
]

for algo, count in algo_counts.items():
    stats_data.append([f'{algo}:', count])

stats_data.extend([
    ['', ''],
    ['DISTRIBUCIÓN POR AÑO', ''],
])

for year, count in year_counts.items():
    stats_data.append([f'{year}:', count])

stats_data.extend([
    ['', ''],
    ['DISTRIBUCIÓN POR TIPO PUBLICACIÓN', ''],
])

for tipo, count in tipo_counts.items():
    stats_data.append([f'{tipo}:', count])

stats_data.extend([
    ['', ''],
    ['DISTRIBUCIÓN POR VALIDACIÓN', ''],
])

for val, count in val_counts.items():
    stats_data.append([f'{val}:', count])

stats_data.extend([
    ['', ''],
    ['DISTRIBUCIÓN POR RELEVANCIA', ''],
])

for rel, count in rel_counts.items():
    stats_data.append([f'{rel}:', count])

stats_data.extend([
    ['', ''],
    ['DISTRIBUCIÓN POR APLICACIÓN', ''],
])

for app, count in app_counts.items():
    stats_data.append([f'{app}:', count])

stats_data.extend([
    ['', ''],
    ['Fecha de actualización:', datetime.now().strftime('%Y-%m-%d %H:%M')],
])

df_stats = pd.DataFrame(stats_data, columns=['ESTADÍSTICAS GENERALES DEL ANÁLISIS', 'Valor'])

# === GUARDAR HOJA ESTADISTICAS ===
wb = load_workbook(ruta)
if 'ESTADISTICAS' in wb.sheetnames:
    del wb['ESTADISTICAS']
wb.save(ruta)

with pd.ExcelWriter(ruta, engine='openpyxl', mode='a') as writer:
    df_stats.to_excel(writer, sheet_name='ESTADISTICAS', index=False)

print(f"✅ ESTADISTICAS actualizado")
print(f"Total papers: {total_papers}")
print(f"Progreso: {progreso:.1f}%")
print(f"Algoritmos: {dict(algo_counts)}")