import pandas as pd
import re
from openpyxl import load_workbook

ruta = r'C:\Users\HangarUPCH\Documents\Antigravity_Proyectos\Swarm Intelligence Algorithms for Multi-RPAS\PAPER_REVISION_SWARM_RPAS\03_ANALISIS_NOTAS\Fichas_Analisis_NUEVO.xlsx'
df = pd.read_excel(ruta, sheet_name='TODOS_PAPERS')

print("🔍 Extrayendo métricas cuantitativas de 33 papers...\n")

# Función para extraer SOLO el número (primer grupo de captura)
def extraer_numero(texto):
    if pd.isna(texto) or not isinstance(texto, str):
        return None
    # Capturar solo dígitos y punto decimal
    match = re.search(r'(\d+\.?\d*)\s*(?:s|seg|ms|J|kJ|%|iter|m|km|units)?', texto)
    if match:
        try:
            return float(match.group(1))  # Grupo 1 = solo el número
        except ValueError:
            return None
    return None

# Función para clasificar complejidad
def clasificar_complejidad(texto):
    if pd.isna(texto): return 'No especificada'
    texto = str(texto).lower()
    if 'polinomial' in texto or 'o(n' in texto: return 'Polinomial'
    if 'exponencial' in texto: return 'Exponencial'
    if 'logarítmica' in texto or 'o(log' in texto: return 'Logarítmica'
    if 'lineal' in texto or 'o(n)' in texto: return 'Lineal'
    if 'pseudo-polinomial' in texto: return 'Pseudo-polinomial'
    return 'Cualitativa'

# Crear lista para métricas normalizadas
metricas_data = []

for idx, row in df.iterrows():
    paper_id = row['ID']
    algoritmo = row['Algoritmo Principal']
    
    # Extraer métricas
    tiempo_num = extraer_numero(row.get('Métrica: Tiempo'))
    energia_num = extraer_numero(row.get('Métrica: Energía'))
    convergencia_num = extraer_numero(row.get('Métrica: Convergencia'))
    
    # Clasificar complejidad
    complejidad = clasificar_complejidad(row.get('Complejidad Computacional'))
    
    metricas_data.append({
        'ID': paper_id,
        'Algoritmo': algoritmo,
        'Tiempo (num)': tiempo_num,
        'Energía (num)': energia_num,
        'Convergencia (num)': convergencia_num,
        'Complejidad': complejidad,
        'Tiempo (texto)': str(row.get('Métrica: Tiempo', ''))[:100] if pd.notna(row.get('Métrica: Tiempo')) else '',
        'Energía (texto)': str(row.get('Métrica: Energía', ''))[:100] if pd.notna(row.get('Métrica: Energía')) else '',
        'Convergencia (texto)': str(row.get('Métrica: Convergencia', ''))[:100] if pd.notna(row.get('Métrica: Convergencia')) else '',
    })
    
    print(f"✅ {paper_id}: {algoritmo} - Complejidad: {complejidad}")

# Crear DataFrame de métricas
df_metricas = pd.DataFrame(metricas_data)

# Guardar como nueva hoja
wb = load_workbook(ruta)
if 'METRICAS_COMPARATIVAS' in wb.sheetnames:
    del wb['METRICAS_COMPARATIVAS']
wb.save(ruta)

with pd.ExcelWriter(ruta, engine='openpyxl', mode='a') as writer:
    df_metricas.to_excel(writer, sheet_name='METRICAS_COMPARATIVAS', index=False)

# Mostrar resumen
print(f"\n{'='*60}")
print(f"✅ MÉTRICAS EXTRAÍDAS")
print(f"{'='*60}")
print(f"Papers procesados: {len(df_metricas)}")
print(f"\nValores numéricos extraídos:")
print(f"  - Tiempo: {df_metricas['Tiempo (num)'].notna().sum()} papers")
print(f"  - Energía: {df_metricas['Energía (num)'].notna().sum()} papers")
print(f"  - Convergencia: {df_metricas['Convergencia (num)'].notna().sum()} papers")
print(f"\nDistribución de complejidad:")
print(df_metricas['Complejidad'].value_counts())