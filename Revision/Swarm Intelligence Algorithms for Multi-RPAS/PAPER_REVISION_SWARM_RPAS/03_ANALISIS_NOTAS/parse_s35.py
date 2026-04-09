import openpyxl

file_path = r'C:\Users\HangarUPCH\Documents\Antigravity_Proyectos\Swarm Intelligence Algorithms for Multi-RPAS\PAPER_REVISION_SWARM_RPAS\03_ANALISIS_NOTAS\extraccion_S34_S35_S36.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['S35']

data = {
    'I4:': 'Sí; utiliza PSO.',
    'I4b:': '1 (single-agent).',
    'I5:': 'Sí; planeación de rutas para pulverización de pesticidas.',
    'E3:': 'Sí (Dado que es single-UAV, se activa la exclusión E3).',
    'DECISIÓN FINAL': 'EXCLUIR [E3] (Nota: es un modelo Single-UAV, incumple requisito Multi-UAV)',
    
    'Tipo de algoritmo SI': 'Híbrido',
    'Nombre exacto del algoritmo': 'Improved Particle Swarm Optimization algorithm combined with the A* algorithm (PSO-A*)',
    'Modificaciones al algoritmo base': 'Factor de convergencia no lineal, K-Means para inicialización, mutación de Cauchy.',
    'Tamaño del enjambre': '50',
    'Número de UAVs en el framework': '1',
    'Tipo de validación': 'Solo simulación',
    'Entorno de simulación utilizado': 'No menciona',
    '¿Incorpora factores ambientales dinámicos?': 'No (entorno estático con obstáculos fijos)',
    'Tipo de aplicación agrícola': 'Fumigación (pesticide spraying)',
    'Parámetros agrícolas específicos': 'Dimensiones de celda/campo (700x700m) y 30 puntos fijos de pulverización',
    
    'Tiempo de ejecución / convergencia': 'No reporta valores numéricos de tiempo.',
    'Consumo de energía / batería': 'No reporta',
    'Criterio de convergencia': 'Iteraciones máximas (200)',
    'Comparación con algoritmo baseline': 'Sí (A*, PSO originario)',
    'Otras métricas reportadas': 'Fitness, longitud de trayectoria, suavidad.',
    
    'C1 —': '0.0',
    'C2 —': '0.33',
    'C3 —': '1.0',
    'C4 —': '1.0',
    'Q =': '0.5325',
    'Clasificación DRS': 'Moderate',
    
    'Q1:': 'Y',
    'Q2:': 'Y',
    'Q3:': 'Y',
    'Q4:': 'N',
    'Q5:': 'Y',
    'Número de Y': '4',
    'Clasificación MMAT': 'High',
    
    'Gaps Tecnológicos identificados': 'Necesidad de integrar visión artificial/Deep Learning para extraer puntos de pulverización.',
    'Gaps Prácticos identificados': 'No mencionados.',
    'Gaps Metodológicos identificados': 'Falta método para seleccionar tareas inicialmente; sin validación en campo real.',
    'Gaps Teóricos identificados': 'Planificación NP-Hard, no hay solución óptima absoluta (solo aproximaciones).',
    
    'Estado de extracción': 'COMPLETO',
    'Fecha de extracción': '29/03/2026',
    'Revisor': 'Antigravity / NotebookLM',
    'Notas adicionales': 'Al igual que el S34, NotebookLM sugirió "INCLUIR" pero extrajo que usa "1" UAV. Se excluye categóricamente bajo la regla E3 para garantizar rigurosidad.',
}

for row in range(1, ws.max_row + 1):
    cell_a = ws.cell(row=row, column=1).value
    if cell_a:
        cell_str = str(cell_a).strip()
        for key, val in data.items():
            if cell_str.startswith(key):
                ws.cell(row=row, column=2).value = val
                break

wb.save(file_path)
print("S35 guardado con exito.")
