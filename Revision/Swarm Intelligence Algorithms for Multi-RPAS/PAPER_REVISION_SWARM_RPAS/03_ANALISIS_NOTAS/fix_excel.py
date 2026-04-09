import openpyxl
import os

file_path = r'C:\Users\HangarUPCH\Documents\Antigravity_Proyectos\Swarm Intelligence Algorithms for Multi-RPAS\PAPER_REVISION_SWARM_RPAS\03_ANALISIS_NOTAS\extraccion_S34_S35_S36.xlsx'
wb = openpyxl.load_workbook(file_path)

# Data dictionaries
data_s34 = {
    'I4:': 'Sí; la inteligencia de enjambres (D*) se usa en la fase de planificación local.',
    'I4b:': '1 (single-agent).',
    'I5:': 'Sí; monitoreo de salud, humedad de suelo y aplicación de recursos.',
    'E3:': 'Sí (Dado que es single-UAV, se activa la exclusión E3).',
    'DECISIÓN FINAL': 'EXCLUIR [E3] (Nota: es un modelo Single-UAV en 2D, incumple requisito Multi-UAV)',
    'Tipo de algoritmo SI': 'Híbrido',
    'Nombre exacto del algoritmo': 'Enhanced Genetic Algorithm using Fuzzy Logic (EGA) e Improved D* Algorithm (ID*)',
    'Modificaciones al algoritmo base': 'GA mejorado con difusa; D* mejorado con función dinámica y Bezier.',
    'Tamaño del enjambre': 'No se menciona numéricamente',
    'Número de UAVs en el framework': '1',
    'Tipo de validación': 'Solo simulación',
    'Entorno de simulación utilizado': 'MATLAB',
    '¿Incorpora factores ambientales dinámicos?': 'Sí, modela obstáculos dinámicos e incertidumbre del viento',
    'Tipo de aplicación agrícola': 'Monitoreo y gestión de cultivos',
    'Parámetros agrícolas específicos': 'Escenario genérico (Grid 90x90 km)',
    'Tiempo de ejecución / convergencia': 'Sí (5.54 segundos)',
    'Consumo de energía / batería': 'No reporta valor consumido',
    'Criterio de convergencia': 'Diferencia de calidad menor a umbral predefinido (épsilon)',
    'Comparación con algoritmo baseline': 'Sí (PSO, ACO, GA, DWA, D*, A*, IACO, Q-Learning)',
    'Otras métricas reportadas': 'Conteo de inflexiones (9), exactitud (91.1%), longitud de ruta',
    'C1 —': '0.0',
    'C2 —': '0.67',
    'C3 —': '1.0',
    'C4 —': '0.0',
    'Q =': '0.4175',
    'Clasificación DRS': 'Moderate',
    'Q1:': 'Y',
    'Q2:': 'Y',
    'Q3:': 'Y',
    'Q4:': 'C',
    'Q5:': 'Y',
    'Número de Y': '4',
    'Clasificación MMAT': 'High',
    'Gaps Tecnológicos identificados': 'Necesidad de hardware para procesar ajustes en tiempo real.',
    'Gaps Prácticos identificados': 'Dificultad de operar en entornos reales diversos y dinámicos.',
    'Gaps Metodológicos identificados': 'Mejorar equilibrio entre velocidad de respuesta, precisión y eficiencia.',
    'Gaps Teóricos identificados': 'Modelo limitado a representación en 2D (simplifica el espacio aéreo 3D real).',
    'Estado de extracción': 'COMPLETO',
    'Fecha de extracción': '29/03/2026',
    'Revisor': 'Antigravity / NotebookLM',
    'Notas adicionales': 'NotebookLM sugirió "INCLUIR", pero registró que es para "1" UAV. Apliqué E3 y lo marqué como EXCLUIDO.',
}

data_s35 = {
    'I4:': 'Sí; utiliza PSO.',
    'I4b:': '1 (single-agent).',
    'I5:': 'Sí; planeación de rutas para pulverización de pesticidas.',
    'E3:': 'Sí (Dado que es single-UAV, se activa la exclusión E3).',
    'DECISIÓN FINAL': 'EXCLUIR [E3] (Nota: es un modelo Single-UAV, incumple requisito Multi-UAV)',
    'Tipo de algoritmo SI': 'Híbrido',
    'Nombre exacto del algoritmo': 'Improved Particle Swarm Optimization algorithm combined with the A* algorithm (PSO-A*)',
    'Modificaciones al algoritmo base': 'Factor de convergencia no lineal, K-Means para inicialización, mutación de Cauchy.',
    'Tamaño del enjambre': '50',
    'Número de UAVs en el framework': '1',
    'Tipo de validación': 'Solo simulación',
    'Entorno de simulación utilizado': 'No menciona',
    '¿Incorpora factores ambientales dinámicos?': 'No (entorno estático con obstáculos fijos)',
    'Tipo de aplicación agrícola': 'Fumigación (pesticide spraying)',
    'Parámetros agrícolas específicos': 'Dimensiones de celda/campo (700x700m) y 30 puntos fijos de pulverización',
    'Tiempo de ejecución / convergencia': 'No reporta valores numéricos de tiempo.',
    'Consumo de energía / batería': 'No reporta',
    'Criterio de convergencia': 'Iteraciones máximas (200)',
    'Comparación con algoritmo baseline': 'Sí (A*, PSO originario)',
    'Otras métricas reportadas': 'Fitness, longitud de trayectoria, suavidad.',
    'C1 —': '0.0',
    'C2 —': '0.33',
    'C3 —': '1.0',
    'C4 —': '1.0',
    'Q =': '0.5325',
    'Clasificación DRS': 'Moderate',
    'Q1:': 'Y',
    'Q2:': 'Y',
    'Q3:': 'Y',
    'Q4:': 'N',
    'Q5:': 'Y',
    'Número de Y': '4',
    'Clasificación MMAT': 'High',
    'Gaps Tecnológicos identificados': 'Necesidad de integrar visión artificial/Deep Learning.',
    'Gaps Prácticos identificados': 'No mencionados.',
    'Gaps Metodológicos identificados': 'Falta método para seleccionar tareas inicialmente; sin validación en campo real.',
    'Gaps Teóricos identificados': 'Planificación NP-Hard, no hay solución óptima absoluta (solo aproximaciones).',
    'Estado de extracción': 'COMPLETO',
    'Fecha de extracción': '29/03/2026',
    'Revisor': 'Antigravity / NotebookLM',
    'Notas adicionales': 'NotebookLM sugirió "INCLUIR" pero extrajo que usa "1" UAV. Se excluye categóricamente bajo la regla E3.',
}

data_s36 = {
    'I4:': 'Sí; emplea versión mejorada de ACO (Inteligencia de Enjambres).',
    'I4b:': '1 (single-agent).',
    'I5:': 'Sí; navegación autónoma en huertos para tareas.',
    'E3:': 'Sí (Dado que es single-UAV en simulación, se activa la exclusión E3).',
    'DECISIÓN FINAL': 'EXCLUIR [E3] (Nota: modelo Single-UAV, incumple el requisito multi-UAV)',
    'Tipo de algoritmo SI': 'ACO (Híbrido con teoría de liderazgo)',
    'Nombre exacto del algoritmo': 'Leader Ant Optimization (LAO) algorithm',
    'Modificaciones al algoritmo base': 'Teoría del líder (bellwether); ajusta feromonas y heurística con parámetros 3D.',
    'Tamaño del enjambre': '50',
    'Número de UAVs en el framework': '1',
    'Tipo de validación': 'Solo simulación',
    'Entorno de simulación utilizado': 'MATLAB R2020a',
    '¿Incorpora factores ambientales dinámicos?': 'No (simulación estática de huerto, no modelan viento/lluvia)',
    'Tipo de aplicación agrícola': 'Monitoreo, fumigación y asistencia en cosecha',
    'Parámetros agrícolas específicos': 'Dimensiones de árboles frutales (5-10m), cercas (1.2-5m).',
    'Tiempo de ejecución / convergencia': 'Sí (0.1021 s promedio)',
    'Consumo de energía / batería': 'No reporta (solo distancia total: 53.74m)',
    'Criterio de convergencia': 'Máximo 100 iteraciones',
    'Comparación con algoritmo baseline': 'Sí (GA, PSO, ACO clásico)',
    'Otras métricas reportadas': 'Fitness medio (44.84), medias de convergencia y p-values.',
    'C1 —': '0.0',
    'C2 —': '0.67',
    'C3 —': '1.0',
    'C4 —': '1.0',
    'Q =': '0.5325',
    'Clasificación DRS': 'Moderate',
    'Q1:': 'Y',
    'Q2:': 'Y',
    'Q3:': 'Y',
    'Q4:': 'Y',
    'Q5:': 'Y',
    'Número de Y': '5',
    'Clasificación MMAT': 'High',
    'Gaps Tecnológicos identificados': 'Necesidad de integración con percepción visual 3D para tiempo real.',
    'Gaps Prácticos identificados': 'Dificultad de mantener ruta óptima ante cambios dinámicos rápidos.',
    'Gaps Metodológicos identificados': 'Falta de comparación con métodos estado-del-arte (SOTA) en 3D.',
    'Gaps Teóricos identificados': 'Selecciones pueden ser subóptimas ante cambios no entrenados.',
    'Estado de extracción': 'COMPLETO',
    'Fecha de extracción': '29/03/2026',
    'Revisor': 'Antigravity / NotebookLM',
    'Notas adicionales': 'Igual que S34 y S35, se descarta porque el framework simulado es para un solo dron (1 UAV).'
}

pages_data = {
    'S34': data_s34,
    'S35': data_s35,
    'S36': data_s36
}

for sheet_name, sheet_data in pages_data.items():
    ws = wb[sheet_name]
    # In Excel, Column A is 1, Column B is 2, Column C is 3. 
    # Based on the CSV dump, the question text is in Column B (2) and the target is Column C (3)
    for row in range(1, ws.max_row + 1):
        cell_label = ws.cell(row=row, column=2).value
        if cell_label:
            cell_str = str(cell_label).strip()
            for key, val in sheet_data.items():
                if cell_str.startswith(key):
                    ws.cell(row=row, column=3).value = str(val)
                    break

wb.save(file_path)
print("Todas las hojas han sido llenadas correctamente.")
