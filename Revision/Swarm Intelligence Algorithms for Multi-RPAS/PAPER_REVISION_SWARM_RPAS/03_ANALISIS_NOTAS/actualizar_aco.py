import pandas as pd
from openpyxl import load_workbook

ruta = r'C:\Users\HangarUPCH\Documents\Antigravity_Proyectos\Swarm Intelligence Algorithms for Multi-RPAS\PAPER_REVISION_SWARM_RPAS\03_ANALISIS_NOTAS\Fichas_Analisis_NUEVO.xlsx'

# Cargar TODOS_PAPERS
df = pd.read_excel(ruta, sheet_name='TODOS_PAPERS')

# Filtrar ACO
aco = df[df['Algoritmo Principal'] == 'ACO'].copy()

# Eliminar hoja ACO existente
wb = load_workbook(ruta)
if 'ACO_ALGORITMOS' in wb.sheetnames:
    del wb['ACO_ALGORITMOS']
wb.save(ruta)

# Guardar nueva hoja ACO
with pd.ExcelWriter(ruta, engine='openpyxl', mode='a') as writer:
    aco.to_excel(writer, sheet_name='ACO_ALGORITMOS', index=False)

print(f"✅ ACO actualizado: {len(aco)} papers")
print(f"IDs: {aco['ID'].tolist()}")