import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

ruta = r'C:\Users\HangarUPCH\Documents\Antigravity_Proyectos\Swarm Intelligence Algorithms for Multi-RPAS\PAPER_REVISION_SWARM_RPAS\03_ANALISIS_NOTAS\Fichas_Analisis_NUEVO.xlsx'
df = pd.read_excel(ruta, sheet_name='TODOS_PAPERS')

# Calcular estadísticas
total = len(df)
analizados = len(df[df['Estado Lectura'] != 'Pendiente'])
progreso = (analizados / total * 100) if total > 0 else 0
algo_counts = df['Algoritmo Principal'].value_counts()

# Crear lista de estadísticas
stats_data = []
stats_data.append(['ESTADÍSTICAS GENERALES DEL ANÁLISIS', ''])
stats_data.append(['Total de Papers:', total])
stats_data.append(['Papers Analizados:', analizados])
stats_data.append(['% Progreso:', f'{progreso:.1f}%'])
stats_data.append(['', ''])
stats_data.append(['DISTRIBUCIÓN POR ALGORITMO', ''])

# Agregar cada algoritmo
for algo, count in algo_counts.items():
    stats_data.append([f'{algo}:', count])

stats_data.append(['', ''])
stats_data.append(['Fecha actualización:', datetime.now().strftime('%Y-%m-%d %H:%M')])

# Crear DataFrame
df_stats = pd.DataFrame(stats_data, columns=['ESTADÍSTICAS', 'Valor'])

# Eliminar hoja existente y guardar nueva
wb = load_workbook(ruta)
if 'ESTADISTICAS' in wb.sheetnames:
    del wb['ESTADISTICAS']
wb.save(ruta)

with pd.ExcelWriter(ruta, engine='openpyxl', mode='a') as writer:
    df_stats.to_excel(writer, sheet_name='ESTADISTICAS', index=False)

print(f"✅ ESTADISTICAS actualizado correctamente")
print(f"Total: {total} | Analizados: {analizados} | Progreso: {progreso:.1f}%")
print(f"\nAlgoritmos detectados: {len(algo_counts)}")
for algo, count in algo_counts.items():
    print(f"  {algo}: {count}")