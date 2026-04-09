import pandas as pd
import os
import fitz
import re
from openpyxl import load_workbook

# Rutas
ruta_excel = r'C:\Users\HangarUPCH\Documents\Antigravity_Proyectos\Swarm Intelligence Algorithms for Multi-RPAS\PAPER_REVISION_SWARM_RPAS\03_ANALISIS_NOTAS\Fichas_Analisis_NUEVO.xlsx'
ruta_pdfs = r'C:\Users\HangarUPCH\Documents\Antigravity_Proyectos\Swarm Intelligence Algorithms for Multi-RPAS\PAPER_REVISION_SWARM_RPAS\02_PAPERS_ORGANIZADOS'

# Papers 021-034 con sus carpetas (buscaremos el nombre exacto)
PAPERS_TARGET = {
    'PAPER_021': 'Algoritmos_PSO',
    'PAPER_022': 'Algoritmos_Otros',
    'PAPER_023': 'Algoritmos_Otros',
    'PAPER_024': 'Algoritmos_PSO',
    'PAPER_025': 'Algoritmos_Otros',
    'PAPER_026': 'Algoritmos_Otros',
    'PAPER_027': 'Algoritmos_Otros',
    'PAPER_028': 'Algoritmos_PSO',
    'PAPER_029': 'Algoritmos_PSO',
    'PAPER_030': 'Algoritmos_PSO',
    'PAPER_031': 'Algoritmos_Otros',
    'PAPER_033': 'Revisiones_Existentes',
    'PAPER_034': 'Revisiones_Existentes',
}

# Cargar Excel
df_metricas = pd.read_excel(ruta_excel, sheet_name='METRICAS_COMPARATIVAS')

print("🔍 Extrayendo métricas de papers 021-034...\n")

def extraer_texto_pdf(ruta_pdf, paginas_max=10):
    try:
        pdf = fitz.open(ruta_pdf)
        texto = ""
        for p in range(min(paginas_max, len(pdf))):
            texto += pdf[p].get_text()
        pdf.close()
        return texto
    except Exception as e:
        return f"ERROR: {str(e)}"

def buscar_metrica(texto, patrones):
    resultados = []
    for patron in patrones:
        matches = re.findall(patron, texto, re.IGNORECASE)
        for match in matches:
            # Si es tupla (por grupos de captura), tomar el primer elemento no vacío
            if isinstance(match, tuple):
                match = [m for m in match if m][0] if any(match) else match[0]
            if match and str(match).strip():
                resultados.append(str(match).strip())
    return resultados[:5]

# Patrones de búsqueda (sin grupos de captura problemáticos)
PATRON_TIEMPO = [
    r'\d+\.?\d*\s*(?:s|sec|seconds|ms|milliseconds|min|minutes)',
    r'time[:\s]+\d+\.?\d*',
    r'execution[:\s]+\d+\.?\d*',
]

PATRON_ENERGIA = [
    r'\d+\.?\d*\s*(?:J|Joules|kJ|Wh|mAh|%)',
    r'energy[:\s]+\d+\.?\d*',
    r'consumption[:\s]+\d+\.?\d*',
    r'distance[:\s]+\d+\.?\d*\s*(?:m|km)',
]

PATRON_CONVERGENCIA = [
    r'\d+\s*(?:iterations|iter|generations|gen)',
    r'convergence[:\s]+\d+\.?\d*',
    r'after[:\s]+\d+\s*iterations',
]

# Procesar cada paper
resultados = []

for paper_id, carpeta in PAPERS_TARGET.items():
    ruta_carpeta = os.path.join(ruta_pdfs, carpeta)
    
    print(f"📄 {paper_id}: ", end="")
    
    # Buscar archivo PDF que empiece con el ID
    pdf_encontrado = None
    if os.path.exists(ruta_carpeta):
        for archivo in os.listdir(ruta_carpeta):
            if archivo.startswith(paper_id) and archivo.endswith('.pdf'):
                pdf_encontrado = archivo
                break
    
    if not pdf_encontrado:
        print(f"❌ PDF no encontrado en {carpeta}")
        resultados.append({'ID': paper_id, 'Estado': 'PDF no encontrado', 'Tiempo': '', 'Energía': '', 'Convergencia': ''})
        continue
    
    ruta_pdf = os.path.join(ruta_carpeta, pdf_encontrado)
    
    # Extraer texto
    texto = extraer_texto_pdf(ruta_pdf, paginas_max=10)
    
    if texto.startswith("ERROR"):
        print(f"❌ Error leyendo PDF")
        resultados.append({'ID': paper_id, 'Estado': 'Error lectura', 'Tiempo': '', 'Energía': '', 'Convergencia': ''})
        continue
    
    # Buscar métricas
    tiempos = buscar_metrica(texto, PATRON_TIEMPO)
    energias = buscar_metrica(texto, PATRON_ENERGIA)
    convergencias = buscar_metrica(texto, PATRON_CONVERGENCIA)
    
    # Formatear resultados
    tiempo_str = '; '.join(tiempos) if tiempos else 'No encontrado'
    energia_str = '; '.join(energias) if energias else 'No encontrado'
    convergencia_str = '; '.join(convergencias) if convergencias else 'No encontrado'
    
    # Extraer primer valor numérico
    tiempo_num = re.search(r'(\d+\.?\d*)', tiempos[0]) if tiempos else None
    energia_num = re.search(r'(\d+\.?\d*)', energias[0]) if energias else None
    convergencia_num = re.search(r'(\d+\.?\d*)', convergencias[0]) if convergencias else None
    
    print(f"✅ T:{len(tiempos)} E:{len(energias)} C:{len(convergencias)}")
    
    resultados.append({
        'ID': paper_id,
        'Estado': 'Procesado',
        'Archivo': pdf_encontrado,
        'Tiempo (encontrado)': tiempo_str[:200],
        'Energía (encontrado)': energia_str[:200],
        'Convergencia (encontrado)': convergencia_str[:200],
        'Tiempo (num)': float(tiempo_num.group(1)) if tiempo_num else None,
        'Energía (num)': float(energia_num.group(1)) if energia_num else None,
        'Convergencia (num)': float(convergencia_num.group(1)) if convergencia_num else None,
    })

# Crear DataFrame
df_resultados = pd.DataFrame(resultados)

# Guardar como nueva hoja
wb = load_workbook(ruta_excel)
if 'METRICAS_021_034' in wb.sheetnames:
    del wb['METRICAS_021_034']
wb.save(ruta_excel)

with pd.ExcelWriter(ruta_excel, engine='openpyxl', mode='a') as writer:
    df_resultados.to_excel(writer, sheet_name='METRICAS_021_034', index=False)

# Mostrar resumen
print(f"\n{'='*60}")
print(f"✅ EXTRACCIÓN COMPLETADA")
print(f"{'='*60}")
print(f"Papers procesados: {len(df_resultados)}")
print(f"\nMétricas encontradas:")
print(f"  - Tiempo: {sum(1 for r in resultados if r['Tiempo (num)'] is not None)} papers")
print(f"  - Energía: {sum(1 for r in resultados if r['Energía (num)'] is not None)} papers")
print(f"  - Convergencia: {sum(1 for r in resultados if r['Convergencia (num)'] is not None)} papers")