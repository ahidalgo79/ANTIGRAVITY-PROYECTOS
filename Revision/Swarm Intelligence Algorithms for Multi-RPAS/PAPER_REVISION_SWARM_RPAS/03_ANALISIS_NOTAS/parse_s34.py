import openpyxl

file_path = r'C:\Users\HangarUPCH\Documents\Antigravity_Proyectos\Swarm Intelligence Algorithms for Multi-RPAS\PAPER_REVISION_SWARM_RPAS\03_ANALISIS_NOTAS\extraccion_S34_S35_S36.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['S34']

data = {
    'I4:': 'Sí; la inteligencia de enjambres (D*) se usa en la fase de planificación local.',
    'I4b:': '1 (single-agent).',
    'I5:': 'Sí; monitoreo de salud, humedad de suelo y aplicación de recursos.',
    'E3:': 'Sí (Dado que es single-UAV, se activa la exclusión E3).',
    'DECISIÓN FINAL': 'EXCLUIR [E3] (Nota: es un modelo Single-UAV en 2D, incumple requisito Multi-UAV)',
    
    'Tipo de algoritmo SI': 'Híbrido',
    'Nombre exacto del algoritmo': 'Enhanced Genetic Algorithm using Fuzzy Logic (EGA) e Improved D* Algorithm (ID*)',
    'Modificaciones al algoritmo base': 'GA mejorado con difusa; D* mejorado con función de costo dinámica y curvas Bezier.',
    'Tamaño del enjambre': 'No se menciona numéricamente',
    'Número de UAVs en el framework': '1',
    'Tipo de validación': 'Solo simulación',
    'Entorno de simulación utilizado': 'MATLAB',
    '¿Incorpora factores ambientales dinámicos?': 'Sí, modela obstáculos dinámicos e incertidumbre del viento',
    'Tipo de aplicación agrícola': 'Monitoreo y gestión de cultivos',
    'Parámetros agrícolas específicos': 'Escenario genérico (Grid 90x90 km)',
    
    'Tiempo de ejecución / convergencia': 'Sí (5.54 segundos)',
    'Consumo de energía / batería': 'No reporta valor consumido (solo lo usa en la función de costo)',
    'Criterio de convergencia': 'Diferencia de calidad menor a umbral predefinido (épsilon)',
    'Comparación con algoritmo baseline': 'Sí (PSO, ACO, GA, DWA, D*, A*, IACO, Q-Learning)',
    'Otras métricas reportadas': 'Conteo de inflexiones (9), exactitud (91.1%), longitud de ruta',
    
    'C1 —': '0.0',
    'C2 —': '0.67',
    'C3 —': '1.0',
    'C4 —': '0.0',
    'Q =': '0.4175',
    'Clasificación DRS': 'Moderate',
    
    'Q1:': 'Y',
    'Q2:': 'Y',
    'Q3:': 'Y',
    'Q4:': 'C',
    'Q5:': 'Y',
    'Número de Y': '4',
    'Clasificación MMAT': 'High',
    
    'Gaps Tecnológicos identificados': 'Necesidad de hardware para procesar ajustes en tiempo real ante obstáculos dinámicos.',
    'Gaps Prácticos identificados': 'Dificultad de operar de forma segura y eficiente en entornos reales diversos y dinámicos.',
    'Gaps Metodológicos identificados': 'Mejorar equilibrio entre velocidad de respuesta, precisión y eficiencia.',
    'Gaps Teóricos identificados': 'Modelo limitado a representación en 2D (simplifica el espacio aéreo 3D real).',
    
    'Estado de extracción': 'COMPLETO',
    'Fecha de extracción': '29/03/2026',
    'Revisor': 'Antigravity / NotebookLM',
    'Notas adicionales': 'Ojo: NotebookLM sugirió "INCLUIR", pero registró que es para "1" UAV. Apliqué E3 y lo marqué como EXCLUIDO.',
}

for row in range(1, ws.max_row + 1):
    cell_a = ws.cell(row=row, column=1).value
    if cell_a:
        cell_str = str(cell_a).strip()
        for key, val in data.items():
            if cell_str.startswith(key):
                ws.cell(row=row, column=2).value = val

wb.save(file_path)
print("S34 guardado con éxito.")
