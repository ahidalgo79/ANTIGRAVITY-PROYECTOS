import pandas as pd
import os
import fitz
import re
from openpyxl import load_workbook

# Rutas
ruta_excel = r'C:\Users\HangarUPCH\Documents\Antigravity_Proyectos\Swarm Intelligence Algorithms for Multi-RPAS\PAPER_REVISION_SWARM_RPAS\03_ANALISIS_NOTAS\Fichas_Analisis_NUEVO.xlsx'
ruta_pdfs = r'C:\Users\HangarUPCH\Documents\Antigravity_Proyectos\Swarm Intelligence Algorithms for Multi-RPAS\PAPER_REVISION_SWARM_RPAS\02_PAPERS_ORGANIZADOS'

# Papers 021-034 con sus carpetas y nombres de archivo
PAPERS_TARGET = {
    'PAPER_021': ('Algoritmos_PSO', 'PAPER_021 - Distributed 3-D Path Planning for Multi-UAVs with Full Area.pdf'),
    'PAPER_022': ('Algoritmos_Otros', 'PAPER_022 - A Task Allocation Strategy of the UAV Swarm Based on Multi.pdf'),
    'PAPER_023': ('Algoritmos_Otros', 'PAPER_023 - Research on Path Planning Method for Mob.pdf'),
    'PAPER_024': ('Algoritmos_PSO', 'PAPER_024 - Three-Dimensional Path Planning of UAV Based on Improved.pdf'),
    'PAPER_025': ('Algoritmos_Otros', 'PAPER_025 - Multi-Unmanned Aerial Vehicle Path Planning Based on Improved.pdf'),
    'PAPER_026': ('Algoritmos_Otros', 'PAPER_026 - Enhancing Swarm Intelligence for Obstacle Avoidance with Multi.pdf'),
    'PAPER_027': ('Algoritmos_Otros', 'PAPER_027 - A Multi-Strategy Collaborative Grey Wolf Optimization.pdf'),
    'PAPER_028': ('Algoritmos_PSO', 'PAPER_028 - UAV Path Planning Algorithm Based on Imp.pdf'),
    'PAPER_029': ('Algoritmos_PSO', 'PAPER_029 - Improved Particle Swarm Optimization Bas.pdf'),
    'PAPER_030': ('Algoritmos_PSO', 'PAPER_030 - Hybrid APF–PSO Algorithm for Regional Dy.pdf'),
    'PAPER_031': ('Algoritmos_Otros', 'PAPER_031 - Three-Dimensional Path Planning of UAV B.pdf'),
    'PAPER_033': ('Revisiones_Existentes', 'PAPER_033 - UAV Formation Trajectory Planning Algorithms.pdf'),
    'PAPER_034': ('Revisiones_Existentes', 'PAPER_034 - Swarm intelligence algorithms for multip.pdf'),
}

# Cargar Excel
df_metricas = pd.read_excel(ruta_excel, sheet_name='METRICAS_COMPARATIVAS')
df_todos = pd.read_excel(ruta_excel, sheet_name='TODOS_PAPERS')

print("🔍 Extrayendo métricas de 14 papers (021-034)...\n")

def extraer_texto_pdf(ruta_pdf, paginas_max=10):
    """Extrae texto de las primeras N páginas del PDF"""
    try:
        pdf = fitz.open(ruta_pdf)
        texto = ""
        for p in range(min(paginas_max, len(pdf))):
            texto += pdf[p].get_text()
        pdf.close()
        return texto
    except Exception as e:
        return f"ERROR: {str(e)}"

def buscar_metrica(texto, patrones):
    """Busca patrones de métricas en el texto"""
    resultados = []
    for patron in patrones:
        matches = re.findall(patron, texto, re.IGNORECASE)
        if matches:
            resultados.extend(matches[:3])  # Máximo 3 coincidencias
    return resultados[:5]  # Máximo 5 resultados totales

# Patrones de búsqueda para cada métrica
PATRON_TIEMPO = [
    r'(\d+\.?\d*)\s*(s|sec|seconds|ms|milliseconds|min|minutes)',
    r'time[:\s]+(\d+\.?\d*)',
    r'execution[:\s]+(\d+\.?\d*)',
    r'computational[:\s]+(\d+\.?\d*)',
]

PATRON_ENERGIA = [
    r'(\d+\.?\d*)\s*(J|Joules|kJ|Wh|mAh|%)',
    r'energy[:\s]+(\d+\.?\d*)',
    r'consumption[:\s]+(\d+\.?\d*)',
    r'battery[:\s]+(\d+\.?\d*)',
    r'distance[:\s]+(\d+\.?\d*)\s*(m|km)',
]

PATRON_CONVERGENCIA = [
    r'(\d+)\s*(iterations|iter|generations|gen)',
    r'convergence[:\s]+(\d+\.?\d*)',
    r'after[:\s]+(\d+)\s*iterations',
    r'within[:\s]+(\d+)\s*iterations',
]

# Procesar cada paper
resultados = []

for paper_id, (carpeta, nombre_archivo) in PAPERS_TARGET.items():
    ruta_pdf = os.path.join(ruta_pdfs, carpeta, nombre_archivo)
    
    print(f"📄 {paper_id}: ", end="")
    
    if not os.path.exists(ruta_pdf):
        print(f"❌ PDF no encontrado")
        resultados.append({'ID': paper_id, 'Estado': 'PDF no encontrado', 'Tiempo': '', 'Energía': '', 'Convergencia': ''})
        continue
    
    # Extraer texto
    texto = extraer_texto_pdf(ruta_pdf, paginas_max=10)
    
    if texto.startswith("ERROR"):
        print(f"❌ Error leyendo PDF")
        resultados.append({'ID': paper_id, 'Estado': 'Error lectura', 'Tiempo': '', 'Energía': '', 'Convergencia': ''})
        continue
    
    # Buscar métricas
    tiempos = buscar_metrica(texto, PATRON_TIEMPO)
    energias = buscar_metrica(texto, PATRON_ENERGIA)
    convergencias = buscar_metrica(texto, PATRON_CONVERGENCIA)
    
    # Formatear resultados
    tiempo_str = '; '.join(tiempos) if tiempos else 'No encontrado'
    energia_str = '; '.join(energias) if energias else 'No encontrado'
    convergencia_str = '; '.join(convergencias) if convergencias else 'No encontrado'
    
    # Extraer primer valor numérico para cada métrica
    tiempo_num = re.search(r'(\d+\.?\d*)', tiempos[0]) if tiempos else None
    energia_num = re.search(r'(\d+\.?\d*)', energias[0]) if energias else None
    convergencia_num = re.search(r'(\d+\.?\d*)', convergencias[0]) if convergencias else None
    
    print(f"✅ T:{len(tiempos)} E:{len(energias)} C:{len(convergencias)}")
    
    resultados.append({
        'ID': paper_id,
        'Estado': 'Procesado',
        'Tiempo (encontrado)': tiempo_str[:200],
        'Energía (encontrado)': energia_str[:200],
        'Convergencia (encontrado)': convergencia_str[:200],
        'Tiempo (num)': float(tiempo_num.group(1)) if tiempo_num else None,
        'Energía (num)': float(energia_num.group(1)) if energia_num else None,
        'Convergencia (num)': float(convergencia_num.group(1)) if convergencia_num else None,
    })

# Crear DataFrame de resultados
df_resultados = pd.DataFrame(resultados)

# Guardar como nueva hoja
wb = load_workbook(ruta_excel)
if 'METRICAS_021_034' in wb.sheetnames:
    del wb['METRICAS_021_034']
wb.save(ruta_excel)

with pd.ExcelWriter(ruta_excel, engine='openpyxl', mode='a') as writer:
    df_resultados.to_excel(writer, sheet_name='METRICAS_021_034', index=False)

# Mostrar resumen
print(f"\n{'='*60}")
print(f"✅ EXTRACCIÓN COMPLETADA")
print(f"{'='*60}")
print(f"Papers procesados: {len(df_resultados)}")
print(f"\nMétricas encontradas:")
print(f"  - Tiempo: {sum(1 for r in resultados if r['Tiempo (num)'] is not None)} papers con valores numéricos")
print(f"  - Energía: {sum(1 for r in resultados if r['Energía (num)'] is not None)} papers con valores numéricos")
print(f"  - Convergencia: {sum(1 for r in resultados if r['Convergencia (num)'] is not None)} papers con valores numéricos")
print(f"\nHoja creada: METRICAS_021_034")