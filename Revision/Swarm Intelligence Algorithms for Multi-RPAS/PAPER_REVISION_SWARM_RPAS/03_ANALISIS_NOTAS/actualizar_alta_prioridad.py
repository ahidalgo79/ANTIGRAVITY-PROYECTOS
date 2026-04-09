import pandas as pd
from openpyxl import load_workbook

ruta = r'C:\Users\HangarUPCH\Documents\Antigravity_Proyectos\Swarm Intelligence Algorithms for Multi-RPAS\PAPER_REVISION_SWARM_RPAS\03_ANALISIS_NOTAS\Fichas_Analisis_NUEVO.xlsx'

# Cargar TODOS_PAPERS
df = pd.read_excel(ruta, sheet_name='TODOS_PAPERS')

# Filtrar por Relevancia = Alta
alta = df[df['Relevancia'] == 'Alta'].copy()

# Eliminar hoja ALTA_PRIORIDAD existente
wb = load_workbook(ruta)
if 'ALTA_PRIORIDAD' in wb.sheetnames:
    del wb['ALTA_PRIORIDAD']
wb.save(ruta)

# Guardar nueva hoja ALTA_PRIORIDAD
with pd.ExcelWriter(ruta, engine='openpyxl', mode='a') as writer:
    alta.to_excel(writer, sheet_name='ALTA_PRIORIDAD', index=False)

print(f"✅ ALTA_PRIORIDAD actualizado: {len(alta)} papers")
print(f"IDs: {alta['ID'].tolist()}")