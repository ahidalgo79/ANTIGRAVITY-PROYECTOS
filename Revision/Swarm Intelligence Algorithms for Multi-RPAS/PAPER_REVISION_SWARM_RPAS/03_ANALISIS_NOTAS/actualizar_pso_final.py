import pandas as pd
from openpyxl import load_workbook

ruta = r'C:\Users\HangarUPCH\Documents\Antigravity_Proyectos\Swarm Intelligence Algorithms for Multi-RPAS\PAPER_REVISION_SWARM_RPAS\03_ANALISIS_NOTAS\Fichas_Analisis_NUEVO.xlsx'

# Cargar TODOS_PAPERS
df = pd.read_excel(ruta, sheet_name='TODOS_PAPERS')

# Filtrar PSO (ahora incluye SPSA que clasificamos como PSO)
pso = df[df['Algoritmo Principal'] == 'PSO'].copy()

# Eliminar hoja PSO_ALGORITMOS existente
wb = load_workbook(ruta)
if 'PSO_ALGORITMOS' in wb.sheetnames:
    del wb['PSO_ALGORITMOS']
wb.save(ruta)

# Guardar nueva hoja PSO_ALGORITMOS
with pd.ExcelWriter(ruta, engine='openpyxl', mode='a') as writer:
    pso.to_excel(writer, sheet_name='PSO_ALGORITMOS', index=False)

print(f"✅ PSO_ALGORITMOS actualizado")
print(f"Total papers PSO: {len(pso)}")
print(f"IDs: {pso['ID'].tolist()}")