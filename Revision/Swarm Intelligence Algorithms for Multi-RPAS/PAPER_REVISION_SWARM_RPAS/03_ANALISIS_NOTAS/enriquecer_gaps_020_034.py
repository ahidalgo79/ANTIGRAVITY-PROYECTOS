import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

ruta = r'C:\Users\HangarUPCH\Documents\Antigravity_Proyectos\Swarm Intelligence Algorithms for Multi-RPAS\PAPER_REVISION_SWARM_RPAS\03_ANALISIS_NOTAS\Fichas_Analisis_NUEVO.xlsx'

# Cargar GAPS_POR_PAPER existente
df_gaps = pd.read_excel(ruta, sheet_name='GAPS_POR_PAPER')

print("📊 Enriqueciendo gaps de papers 020-034 con información del Ideas_Gaps.docx...\n")
print(f"Gaps actuales: {len(df_gaps)}")

# === DATOS ENRIQUECIDOS DESDE IDEAS_GAPS.DOCX (Papers 020-034) ===
# Información extraída de las secciones finales del documento
GAPS_ENRIQUECIDOS_020_034 = {
    # PAPER_020 - ACO Convergence Analysis (MMACO)
    'PAPER_020': [
        {'causa': 'Complejidad logística y recursos; estudio centrado en validación algorítmica mediante simulación numérica',
         'consecuencia': 'No se consideran latencia de sensores, fricción de aire real o fallos mecánicos',
         'sugerencia': 'Implementar despliegue en hardware físico para validación operativa'},
        {'causa': 'Complejidad computacional de algoritmos bioinspirados requiere procesamiento previo extenso',
         'consecuencia': 'UAVs podrían no reaccionar con rapidez necesaria ante obstáculos que aparezcan espontáneamente durante vuelo',
         'sugerencia': 'Optimizar algoritmo para respuesta en tiempo real sin procesamiento previo extenso'},
        {'causa': 'Diseño experimental limitado a analizar comportamiento de solo dos UAVs',
         'consecuencia': 'No se garantiza que algoritmo mantenga eficiencia o evasión de colisiones en enjambres masivos (50+ drones)',
         'sugerencia': 'Evaluar escalabilidad con 5, 10, 25 y 50 UAVs'},
        {'causa': 'Enfoque centrado estrictamente en costo de ruta y tiempo de convergencia',
         'consecuencia': 'No se evalúa si ruta óptima en distancia es también la más eficiente en términos de batería',
         'sugerencia': 'Incorporar métricas de consumo energético y latencia en evaluación'},
    ],
    
    # PAPER_021 - PSO Distributed 3-D Path Planning
    'PAPER_021': [
        {'causa': 'Enfoque en planificación distribuida para múltiples UAVs pero sin validación física',
         'consecuencia': 'No se considera latencia de comunicación real entre drones en entorno de campo',
         'sugerencia': 'Validar con enjambre real de UAVs comunicándose en tiempo real'},
        {'causa': 'Supuesto de información completa del entorno 3D',
         'consecuencia': 'El algoritmo puede fallar en escenarios de exploración donde el mapa se construye incrementalmente',
         'sugerencia': 'Desarrollar variante para información parcial con descubrimiento activo de obstáculos'},
        {'causa': 'Complejidad computacional de coordinación distribuida',
         'consecuencia': 'Tiempo de cálculo puede exceder límites para replanificación en tiempo real',
         'sugerencia': 'Optimizar para hardware embebido con restricciones de CPU y memoria'},
        {'causa': 'Ausencia de modelado de fallos de comunicación',
         'consecuencia': 'El sistema puede colapsar si un nodo pierde conectividad con el enjambre',
         'sugerencia': 'Implementar protocolos tolerantes a fallos de comunicación'},
        {'causa': 'Enfoque en cobertura de área completa sin considerar energía',
         'consecuencia': 'UAVs podrían agotar batería antes de completar la misión de cobertura',
         'sugerencia': 'Integrar modelo de consumo energético en función de costo de cobertura'},
    ],
    
    # PAPER_022 - WPA Task Allocation
    'PAPER_022': [
        {'causa': 'Enfoque en asignación de tareas con Wolf Pack Algorithm sin validación física',
         'consecuencia': 'No se considera overhead de comunicación para coordinación de asignación en tiempo real',
         'sugerencia': 'Medir latencia de asignación en hardware real con múltiples UAVs'},
        {'causa': 'Supuesto de tareas estáticas predefinidas',
         'consecuencia': 'El sistema no puede reasignar dinámicamente si surgen nuevas tareas durante la misión',
         'sugerencia': 'Desarrollar reasignación dinámica de tareas en tiempo de ejecución'},
        {'causa': 'Ausencia de modelado de prioridades de tarea',
         'consecuencia': 'Todas las tareas se tratan igual, sin considerar urgencia o valor estratégico',
         'sugerencia': 'Incorporar sistema de priorización de tareas en función de aptitud'},
        {'causa': 'Validación limitada a simulación numérica',
         'consecuencia': 'No se evalúa robustez ante fallos de UAVs individuales durante ejecución de tareas',
         'sugerencia': 'Implementar pruebas con fallos simulados de agentes durante misión'},
    ],
    
    # PAPER_023 - Hybrid Swarm Multi-Dimensional
    'PAPER_023': [
        {'causa': 'Algoritmo híbrido combina múltiples estrategias pero aumenta complejidad computacional',
         'consecuencia': 'Tiempo de procesamiento puede ser prohibitivo para aplicaciones en tiempo real',
         'sugerencia': 'Analizar trade-off entre calidad de solución y tiempo de cómputo para hardware embebido'},
        {'causa': 'Validación en entornos multi-dimensionales simulados',
         'consecuencia': 'No se considera cómo se traduce a restricciones físicas reales de vuelo',
         'sugerencia': 'Mapear dimensiones algorítmicas a parámetros físicos medibles del UAV'},
        {'causa': 'Ausencia de comparación con estándares de la industria',
         'consecuencia': 'Difícil evaluar superioridad real frente a métodos establecidos como RRT* o A*',
         'sugerencia': 'Incluir benchmarks de robótica clásica en evaluación comparativa'},
        {'causa': 'Enfoque en plataformas móviles sin especificar tipo de vehículo',
         'consecuencia': 'Resultados pueden no ser directamente aplicables a UAVs con restricciones aerodinámicas',
         'sugerencia': 'Especificar restricciones cinemáticas de UAVs en modelo de planificación'},
    ],
    
    # PAPER_024 - PSO 3D Path Planning Improved
    'PAPER_024': [
        {'causa': 'Mejoras al PSO enfocadas en convergencia pero sin validación en hardware',
         'consecuencia': 'No se conoce consumo de recursos para ejecución en procesadores de vuelo',
         'sugerencia': 'Medir uso de CPU, memoria y energía en hardware representativo de UAV'},
        {'causa': 'Entornos 3D modelados con obstáculos estáticos',
         'consecuencia': 'El algoritmo puede fallar ante amenazas móviles o cambios imprevistos del entorno',
         'sugerencia': 'Integrar módulo de replanificación reactiva ante obstáculos dinámicos'},
        {'causa': 'Ausencia de modelado de incertidumbre en posición',
         'consecuencia': 'Errores de GPS o IMU pueden causar desviaciones de la trayectoria planificada',
         'sugerencia': 'Incorporar márgenes de seguridad basados en incertidumbre de localización'},
        {'causa': 'Enfoque en ruta individual sin coordinación multi-UAV',
         'consecuencia': 'No se evalúa evitación de colisiones entre múltiples drones en mismo espacio',
         'sugerencia': 'Extender para planificación coordinada de enjambres con evitación mutua'},
    ],
    
    # PAPER_025 - NOA Nutcracker Optimization
    'PAPER_025': [
        {'causa': 'Algoritmo NOA es emergente con poca validación comparativa en literatura UAV',
         'consecuencia': 'Difícil establecer superioridad real frente a algoritmos más estudiados como PSO o ACO',
         'sugerencia': 'Realizar comparación exhaustiva con 5+ algoritmos establecidos en mismos benchmarks'},
        {'causa': 'Validación exclusiva mediante simulación',
         'consecuencia': 'No se considera ruido de sensores, latencia o limitaciones de hardware embebido',
         'sugerencia': 'Implementar pruebas de campo con UAVs reales en entorno controlado'},
        {'causa': 'Ausencia de análisis de escalabilidad',
         'consecuencia': 'No se conoce comportamiento del NOA con enjambres de 10, 50 o 100 UAVs',
         'sugerencia': 'Evaluar escalabilidad del algoritmo con diferentes tamaños de enjambre'},
        {'causa': 'Enfoque en optimización de ruta sin considerar energía',
         'consecuencia': 'Ruta más corta puede no ser la más eficiente energéticamente para el UAV específico',
         'sugerencia': 'Incorporar modelo de consumo energético específico por tipo de UAV'},
    ],
    
    # PAPER_026 - DBO Dung Beetle Optimization
    'PAPER_026': [
        {'causa': 'DBO es algoritmo bioinspirado reciente con poca aplicación en planificación UAV',
         'consecuencia': 'Falta comprensión de cómo parámetros del algoritmo se relacionan con restricciones de vuelo',
         'sugerencia': 'Desarrollar guía de mapeo entre parámetros DBO y restricciones cinemáticas de UAV'},
        {'causa': 'Validación en entornos con obstáculos estáticos',
         'consecuencia': 'No se evalúa capacidad de evasión de obstáculos móviles en tiempo real',
         'sugerencia': 'Integrar pruebas con obstáculos dinámicos de velocidad variable'},
        {'causa': 'Ausencia de análisis de convergencia en alta dimensionalidad',
         'consecuencia': 'El algoritmo puede estancarse en óptimos locales en espacios de búsqueda complejos',
         'sugerencia': 'Implementar mecanismos de escape de óptimos locales inspirados en otros algoritmos'},
        {'causa': 'Enfoque en navegación de robot móvil sin especificar UAV',
         'consecuencia': 'Resultados pueden no considerar restricciones aerodinámicas de vuelo',
         'sugerencia': 'Adaptar modelo específicamente para dinámica de UAVs'},
    ],
    
    # PAPER_027 - GWO Multi-Strategy Collaborative
    'PAPER_027': [
        {'causa': 'Estrategias múltiples mejoran convergencia pero aumentan complejidad de ajuste',
         'consecuencia': 'Usuarios requieren experiencia significativa para configurar parámetros óptimamente',
         'sugerencia': 'Desarrollar método de ajuste automático de parámetros basado en características del entorno'},
        {'causa': 'Validación mediante simulación numérica en Matlab',
         'consecuencia': 'No se evalúa desempeño en hardware con recursos limitados de procesamiento',
         'sugerencia': 'Probar en procesadores embebidos típicos de UAVs comerciales'},
        {'causa': 'Ausencia de modelado de comunicación entre agentes',
         'consecuencia': 'No se considera cómo latencia o pérdida de paquetes afecta coordinación del enjambre',
         'sugerencia': 'Incorporar modelo de red con latencia y pérdida de comunicación en simulación'},
        {'causa': 'Enfoque en planificación de ruta individual',
         'consecuencia': 'No se evalúa escalabilidad a coordinación de enjambres masivos',
         'sugerencia': 'Extender GWO colaborativo para enjambres de 20+ UAVs'},
    ],
    
    # PAPER_028 - PSO UAV Path Planning Improved
    'PAPER_028': [
        {'causa': 'Mejoras al PSO enfocadas en eficiencia de ruta sin validación física',
         'consecuencia': 'No se conoce viabilidad de implementación en controladores de vuelo comerciales',
         'sugerencia': 'Validar en hardware de vuelo real con medición de recursos'},
        {'causa': 'Entornos de prueba con obstáculos estáticos modelados geométricamente',
         'consecuencia': 'El algoritmo puede no manejar eficazmente obstáculos con formas irregulares reales',
         'sugerencia': 'Probar con mapas de obstáculos derivados de datos LiDAR o fotogrametría real'},
        {'causa': 'Ausencia de análisis de robustez ante fallos',
         'consecuencia': 'No se evalúa comportamiento del algoritmo si UAV pierde localización o comunicación',
         'sugerencia': 'Implementar pruebas de robustez con fallos simulados de sensores'},
        {'causa': 'Enfoque en optimización de trayectoria sin considerar viento',
         'consecuencia': 'Ruta óptima en condiciones calmadas puede ser ineficiente con viento en contra',
         'sugerencia': 'Incorporar modelo de viento en función de costo de trayectoria'},
    ],
    
    # PAPER_029 - PSO Fuzzy Controller Fusion
    'PAPER_029': [
        {'causa': 'Fusión con controlador difuso añade complejidad computacional significativa',
         'consecuencia': 'Tiempo de inferencia difusa puede exceder límites de tiempo real para control de vuelo',
         'sugerencia': 'Optimizar reglas difusas para reducir tiempo de inferencia en hardware embebido'},
        {'causa': 'Validación mediante simulación de control de trayectoria',
         'consecuencia': 'No se evalúa desempeño con perturbaciones reales como ráfagas de viento',
         'sugerencia': 'Implementar pruebas con perturbaciones ambientales controladas en campo'},
        {'causa': 'Ausencia de análisis de estabilidad teórica',
         'consecuencia': 'No hay garantías matemáticas de que el sistema difuso-PSO sea estable en todas las condiciones',
         'sugerencia': 'Desarrollar análisis de estabilidad Lyapunov para el sistema híbrido'},
        {'causa': 'Enfoque en seguimiento de trayectoria sin planificación de ruta',
         'consecuencia': 'El sistema asume ruta predefinida, no genera rutas nuevas ante obstáculos',
         'sugerencia': 'Integrar planificación de ruta global con seguimiento local difuso'},
    ],
    
    # PAPER_030 - Hybrid APF-PSO Formation
    'PAPER_030': [
        {'causa': 'Formación dinámica de enjambre modelada como masas puntuales sin dinámica real',
         'consecuencia': 'Las trayectorias de formación pueden no ser ejecutables por UAVs con inercia y limitaciones de giro',
         'sugerencia': 'Incorporar modelo dinámico completo de UAV en simulación de formación'},
        {'causa': 'Supuesto de amenaza con movimiento lineal uniforme',
         'consecuencia': 'La barrera defensiva puede fallar ante amenazas con maniobras evasivas complejas',
         'sugerencia': 'Modelar amenazas con comportamiento evasivo inteligente y no lineal'},
        {'causa': 'Ausencia de pruebas de campo con enjambre real',
         'consecuencia': 'No se evalúa sincronización real de comunicación para mantenimiento de formación',
         'sugerencia': 'Implementar pruebas con 5+ UAVs reales manteniendo formación dinámica'},
        {'causa': 'Enfoque en barrera planar 2D para defensa',
         'consecuencia': 'La cobertura defensiva es limitada en escenarios 3D con amenazas aéreas',
         'sugerencia': 'Desarrollar formación defensiva multicapa en espacio 3D'},
        {'causa': 'Escalabilidad probada hasta 25 drones sin protocolos distribuidos',
         'consecuencia': 'No se garantiza coordinación efectiva para enjambres masivos de 50+ UAVs',
         'sugerencia': 'Desarrollar protocolos de consenso distribuido para enjambres masivos'},
    ],
    
    # PAPER_031 - DOA Dream Optimization
    'PAPER_031': [
        {'causa': 'DOA es algoritmo muy reciente con poca validación en aplicaciones UAV',
         'consecuencia': 'Falta comprensión de ventajas y desventajas frente a algoritmos establecidos',
         'sugerencia': 'Realizar estudio comparativo exhaustivo con 10+ algoritmos en benchmarks estándar'},
        {'causa': 'Validación exclusiva mediante simulación numérica',
         'consecuencia': 'No se considera viabilidad de implementación en hardware de vuelo real',
         'sugerencia': 'Implementar en procesador embebido y medir consumo de recursos'},
        {'causa': 'Ausencia de análisis de escalabilidad a multi-UAV',
         'consecuencia': 'No se conoce comportamiento del DOA en coordinación de enjambres',
         'sugerencia': 'Extender DOA para planificación coordinada de múltiples agentes'},
        {'causa': 'Enfoque en optimización 3D sin restricciones cinemáticas',
         'consecuencia': 'Rutas generadas pueden requerir maniobras físicamente imposibles para el UAV',
         'sugerencia': 'Incorporar restricciones de radio de giro y velocidad en modelo de optimización'},
    ],
    
    # PAPER_033 - Review Formation Trajectory
    'PAPER_033': [
        {'causa': 'Artículo de revisión identifica gaps pero no los valida experimentalmente',
         'consecuencia': 'Los gaps identificados permanecen como observaciones teóricas sin cuantificación',
         'sugerencia': 'Seleccionar 3-5 gaps críticos y validarlos experimentalmente en investigación futura'},
        {'causa': 'Revisión cubre literatura hasta fecha de publicación sin estudios más recientes',
         'consecuencia': 'Algoritmos emergentes (2024-2025) no están incluidos en el análisis',
         'sugerencia': 'Actualizar revisión sistemáticamente cada 2 años con nuevos estudios'},
        {'causa': 'Enfoque en formación de trayectoria sin integración con planificación de ruta',
         'consecuencia': 'Existe brecha entre planificación global y control de formación local',
         'sugerencia': 'Desarrollar framework unificado que integre planificación y control de formación'},
        {'causa': 'Ausencia de benchmarks estandarizados para comparación de algoritmos de formación',
         'consecuencia': 'Difícil comparar objetivamente diferentes enfoques de planificación de formación',
         'sugerencia': 'Establecer benchmarks públicos con métricas estandarizadas para formación de UAVs'},
    ],
    
    # PAPER_034 - Review SI for Multi-UAV
    'PAPER_034': [
        {'causa': 'Revisión comprehensiva identifica tendencias pero no valida experimentalmente',
         'consecuencia': 'Recomendaciones permanecen teóricas sin evidencia empírica de efectividad',
         'sugerencia': 'Implementar estudio experimental que valide las recomendaciones de la revisión'},
        {'causa': 'Enfoque en colaboración multi-UAV sin considerar heterogeneidad de plataformas',
         'consecuencia': 'Resultados pueden no aplicar a flotas con UAVs de diferentes capacidades',
         'sugerencia': 'Incluir análisis de colaboración en flotas heterogéneas (ala fija + multirrotor)'},
        {'causa': 'Ausencia de análisis de aspectos regulatorios y de espacio aéreo',
         'consecuencia': 'Barreras legales para implementación de enjambres no se abordan en la revisión',
         'sugerencia': 'Incorporar análisis de marcos regulatorios para operaciones de enjambre UAV'},
        {'causa': 'Revisión cubre algoritmos de inteligencia de enjambre sin métodos emergentes de IA',
         'consecuencia': 'Técnicas como Deep RL o GNN no se comparan con métodos de enjambre tradicionales',
         'sugerencia': 'Expandir revisión para incluir métodos de aprendizaje profundo aplicados a enjambres'},
    ],
}

# === CONTADORES ===
gaps_enriquecidos = 0
gaps_sin_cambios = 0

# === VERIFICAR COLUMNAS ===
columnas_nuevas = ['Causa Raíz', 'Consecuencia', 'Sugerencia']
for col in columnas_nuevas:
    if col not in df_gaps.columns:
        df_gaps[col] = ''

# === ACTUALIZAR GAPS DE PAPERS 020-034 ===
for idx, row in df_gaps.iterrows():
    paper_id = row['ID']
    
    # Solo actualizar papers 020-034
    if paper_id in GAPS_ENRIQUECIDOS_020_034:
        lista_gaps = GAPS_ENRIQUECIDOS_020_034[paper_id]
        
        # Encontrar gaps de este paper en el DataFrame
        paper_gaps = df_gaps[df_gaps['ID'] == paper_id].index.tolist()
        
        # Asignar cíclicamente si hay más gaps en Excel que en datos enriquecidos
        for i, gap_idx in enumerate(paper_gaps):
            if i < len(lista_gaps):
                gap_data = lista_gaps[i]
                df_gaps.at[gap_idx, 'Causa Raíz'] = gap_data['causa']
                df_gaps.at[gap_idx, 'Consecuencia'] = gap_data['consecuencia']
                df_gaps.at[gap_idx, 'Sugerencia'] = gap_data['sugerencia']
                gaps_enriquecidos += 1
            else:
                gaps_sin_cambios += 1

# === GUARDAR EXCEL ACTUALIZADO ===
wb = load_workbook(ruta)
if 'GAPS_POR_PAPER' in wb.sheetnames:
    del wb['GAPS_POR_PAPER']
wb.save(ruta)

with pd.ExcelWriter(ruta, engine='openpyxl', mode='a') as writer:
    df_gaps.to_excel(writer, sheet_name='GAPS_POR_PAPER', index=False)

# === MOSTRAR RESULTADOS ===
print(f"\n{'='*60}")
print(f"✅ GAPS DE PAPERS 020-034 ENRIQUECIDOS EXITOSAMENTE")
print(f"{'='*60}")
print(f"\nTotal de gaps: {len(df_gaps)}")
print(f"Gaps enriquecidos (020-034): {gaps_enriquecidos}")
print(f"Gaps previos (001-019): 110")
print(f"Gaps TOTALES enriquecidos: {gaps_enriquecidos + 110}")
print(f"\nDistribución por Paper (020-034):")
for paper_id in sorted(GAPS_ENRIQUECIDOS_020_034.keys()):
    count = len(df_gaps[df_gaps['ID'] == paper_id])
    print(f"  {paper_id}: {count} gaps")

print(f"\nColumnas finales: {df_gaps.columns.tolist()}")
print(f"\nEjemplo (PAPER_030):")
print(df_gaps[df_gaps['ID'] == 'PAPER_030'][['ID', 'Dimensión', 'Gap', 'Causa Raíz', 'Consecuencia', 'Sugerencia']].head(2).to_string())