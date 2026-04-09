import openpyxl

file_path = r'C:\Users\HangarUPCH\Documents\Antigravity_Proyectos\Swarm Intelligence Algorithms for Multi-RPAS\PAPER_REVISION_SWARM_RPAS\04_BIBLIOGRAFIA\Rescreening_TA_DrGarza.xlsx'
wb = openpyxl.load_workbook(file_path)
ws = wb['SCREENING_SAMPLE']

# Col H = DECISION, Col I = CRITERIO, Col J = NOTAS
# Data rows start at row 4 (rows 1-3 are headers/instructions)
# Format: (ID, Decision, Criterio, Notas)
decisions = [
    # TA_001
    ("TA_001", "EXCLUIR", "E4",
     "Multi-agent Reinforcement Learning (MARL) — no es un algoritmo SI clásico (PSO/ACO/ABC/GWO). Sin abstract para verificar hibridación con SI."),
    # TA_002
    ("TA_002", "EXCLUIR", "E4",
     "Multi-UAV agrícola y aplicación relevante, pero el abstract no menciona ningún algoritmo SI. Menciona restricciones energéticas y cobertura pero el método no es identificable como SI."),
    # TA_003
    ("TA_003", "EXCLUIR", "E5",
     "Planificación de rutas para redes de cables eléctricos urbanas. No es UAV ni dominio agrícola."),
    # TA_004
    ("TA_004", "EXCLUIR", "E5",
     "Robot de poda de jardines (terrestre). Aunque usa ACO-Genético (SI), no es UAV. Dominio no agrícola en el contexto de la revisión."),
    # TA_005
    ("TA_005", "EXCLUIR", "E5",
     "Robot móvil en entorno radiactivo (planta nuclear). No es UAV ni dominio agrícola."),
    # TA_006
    ("TA_006", "EXCLUIR", "E5",
     "Sistema híbrido para monitoreo y limpieza de vehículos no tripulados. Sin enfoque en path planning agrícola con SI."),
    # TA_007
    ("TA_007", "EXCLUIR", "E5",
     "Robot móvil terrestre con PSO + curvas Bezier. No es UAV ni aplicación agrícola."),
    # TA_008
    ("TA_008", "EXCLUIR", "E5",
     "Robot de patrulla en planta nuclear. PSO multi-objetivo, pero dominio industrial/nuclear. No es UAV agrícola."),
    # TA_009
    ("TA_009", "EXCLUIR", "E5",
     "Sistema de gestión de drones para ciudades inteligentes. No es path planning agrícola multi-UAV con SI."),
    # TA_010
    ("TA_010", "EXCLUIR", "E5",
     "PSO jerárquico para multi-UAV bajo 'varios tipos de amenazas'. Contexto militar/defensa, no agrícola."),
    # TA_011
    ("TA_011", "EXCLUIR", "E5",
     "Algoritmo de path planning para robot con objetivo dinámico. No es UAV ni dominio agrícola."),
    # TA_012
    ("TA_012", "EXCLUIR", "E4",
     "Cobertura multi-UAV con algoritmo de subasta (auction). No es SI (PSO/ACO/etc.). Entornos dinámicos genéricos."),
    # TA_013
    ("TA_013", "EXCLUIR", "E4",
     "UAV en misiones de muestreo en agricultura de precisión — dominio relevante. Sin abstract disponible; el término 'routing' sugiere VRP/TSP, no SI. Requiere verificación manual del texto completo."),
    # TA_014
    ("TA_014", "EXCLUIR", "E2",
     "Título 'Computations and Methods' completamente genérico. Sin abstract. No se puede evaluar. Posible entrada bibliográfica incompleta o mal catalogada."),
    # TA_015
    ("TA_015", "EXCLUIR", "E5",
     "UAV en 'entornos complejos'. Sin abstract; el título no indica aplicación agrícola ni SI específico."),
    # TA_016
    ("TA_016", "EXCLUIR", "E5",
     "Vigilancia/surveillance con UAV (enfoque de seguridad o reconocimiento). Método 'greedy' distribuido, no SI. Dominio no agrícola."),
    # TA_017
    ("TA_017", "EXCLUIR", "E5",
     "Maquinaria agrícola terrestre (grupos de maquinarias). No es UAV. Aunque es multi-agente y agrícola, no aplica al dominio de la revisión."),
    # TA_018
    ("TA_018", "EXCLUIR", "E5",
     "Metodología de optimización genérica para path planning con múltiples restricciones. Sin mención de UAV ni agricultura."),
    # TA_019
    ("TA_019", "EXCLUIR", "E4",
     "Deep Reinforcement Learning para recolección de datos IoT. No es SI y el dominio no es agrícola."),
    # TA_020
    ("TA_020", "EXCLUIR", "E5",
     "Multi-UAV 3D path planning con algoritmo híbrido. Potencialmente relevante en método, pero sin aplicación agrícola identificada en título ni abstract."),
    # TA_021
    ("TA_021", "EXCLUIR", "E4",
     "UAV en manzanares (dominio agrícola relevante). Usa Digital Twin + grafo ponderado. El enfoque es de grafo/routing, no SI clásico. Sin hibridación SI confirmada."),
    # TA_022
    ("TA_022", "EXCLUIR", "E5",
     "Inspección óptica automatizada en manufactura. No es UAV ni dominio agrícola."),
    # TA_023
    ("TA_023", "EXCLUIR", "E5",
     "Revisión de robótica agrícola para reducción de carbono. No es path planning de UAV con SI. Revisión general del área."),
    # TA_024
    ("TA_024", "EXCLUIR", "E5",
     "Encuesta sobre 'Age of Information' en redes IoT masivas. No es UAV path planning ni dominio agrícola."),
    # TA_025
    ("TA_025", "EXCLUIR", "E5",
     "UAV Swarm SAR (Synthetic Aperture Radar). Formación de imágenes radar, no path planning. No es dominio agrícola."),
    # TA_026
    ("TA_026", "EXCLUIR", "E5",
     "Revisión sobre descarga computacional (offloading) en redes de enjambres UAV. No es path planning con SI agrícola."),
    # TA_027
    ("TA_027", "EXCLUIR", "E5",
     "ACO mejorado con Monte Carlo para robot soldador. SI (ACO) + path planning, pero robot industrial (no UAV, no agrícola)."),
    # TA_028
    ("TA_028", "EXCLUIR", "E5",
     "Revisión sobre ML en cadena de suministro agrícola. No es path planning de UAV."),
    # TA_029
    ("TA_029", "EXCLUIR", "E5",
     "ACO mejorado para evacuación de emergencia en edificios. SI + path planning, pero dominio de seguridad civil (no UAV, no agrícola)."),
    # TA_030
    ("TA_030", "EXCLUIR", "E5",
     "Path planning integrado para pelotones de vehículos autónomos terrestres. No es UAV ni dominio agrícola."),
    # TA_031
    ("TA_031", "EXCLUIR", "E5",
     "Robot terrestre de invernadero con Dung Beetle Algorithm (SI). Agricultura + SI, pero robot terrestre (no UAV). Potencialmente interesante para marco comparativo."),
    # TA_032
    ("TA_032", "EXCLUIR", "E5",
     "IoRT + AI para fumigación de herbicidas. Dominio agrícola, pero sin confirmación de multi-UAV con SI. Posiblemente robot terrestre IoT."),
    # TA_033
    ("TA_033", "EXCLUIR", "E5",
     "Navegación indoor para personas con discapacidad visual. Sin relación con UAV ni agricultura."),
    # TA_034
    ("TA_034", "EXCLUIR", "E5",
     "Identificación de parámetros modales con mediciones de UAV (sensado estructural). No es path planning ni dominio agrícola."),
    # TA_035
    ("TA_035", "EXCLUIR", "E5",
     "Seguridad en redes IoT habilitadas por drones (Machine Learning). No es path planning con SI ni dominio agrícola."),
    # TA_036
    ("TA_036", "EXCLUIR", "E5",
     "Robot móvil en ambiente radiactivo indoor (bi-level híbrido). SI potencial + path planning, pero no UAV y no agrícola."),
    # TA_037
    ("TA_037", "EXCLUIR", "E5",
     "UAV 3D path planning con NGO (metaheurística SI). 'Multiple threats' indica contexto militar/defensa. No es dominio agrícola."),
    # TA_038
    ("TA_038", "EXCLUIR", "E5",
     "Vehículos de superficie no tripulados (USV/marítimos). No es UAV ni dominio agrícola."),
    # TA_039
    ("TA_039", "EXCLUIR", "E5",
     "Multi-robot terrestre con Q-learning + PSO para transporte de objetos. Híbrido SI, pero dominio industrial (no UAV, no agrícola)."),
    # TA_040
    ("TA_040", "EXCLUIR", "E5",
     "Vehículos terrestres autónomos con PSO adaptativo. No es UAV ni dominio agrícola."),
    # TA_041
    ("TA_041", "EXCLUIR", "E5",
     "MAHACO: ACO híbrido multi-algoritmo para grupo de UAVs en 3D. Técnicamente relevante (multi-UAV + SI + 3D). Sin embargo, abstract no menciona aplicación agrícola — dominio genérico. NOTA: Caso borderline; revisar texto completo para descartar agricultura."),
    # TA_042
    ("TA_042", "EXCLUIR", "E5",
     "Revisión general sobre UAVs (aspectos prácticos, seguridad, tendencias). No es path planning específico con SI en agricultura."),
    # TA_043
    ("TA_043", "EXCLUIR", "E5",
     "SI para coordinación multi-robot en entornos industriales. SI + multi-agente, pero dominio industrial (no UAV, no agrícola)."),
    # TA_044
    ("TA_044", "EXCLUIR", "E5",
     "Taxonomía y análisis de algoritmos metaheurísticos (revisión general desde 2020). No es path planning UAV agrícola específico."),
]

# Find rows by ID in column B and fill H, I, J
id_to_row = {}
for row in range(1, ws.max_row + 1):
    val = ws.cell(row=row, column=2).value
    if val and str(val).startswith("TA_"):
        id_to_row[str(val).strip()] = row

print(f"IDs encontrados: {len(id_to_row)}")

for rec_id, decision, criterio, notas in decisions:
    if rec_id in id_to_row:
        r = id_to_row[rec_id]
        ws.cell(row=r, column=8).value = decision   # H: DECISIÓN_REVISOR2
        ws.cell(row=r, column=9).value = criterio   # I: CRITERIO
        ws.cell(row=r, column=10).value = notas     # J: NOTAS_REVISOR2
        print(f"  {rec_id}: {decision} [{criterio}]")
    else:
        print(f"  ADVERTENCIA: {rec_id} no encontrado en la hoja")

wb.save(file_path)
print(f"\nArchivo guardado correctamente: {file_path}")
print(f"\nRESUMEN: {len(decisions)} registros evaluados.")
incluidos = sum(1 for _, d, _, _ in decisions if d == "INCLUIR")
excluidos = sum(1 for _, d, _, _ in decisions if d == "EXCLUIR")
print(f"  INCLUIR: {incluidos}")
print(f"  EXCLUIR: {excluidos}")
